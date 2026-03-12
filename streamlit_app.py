"""
V Squared Sector Search Analysis — Streamlit Web App

Features:
- Multiple concurrent searches, each in its own tab
- Runs persist on GitHub Actions — closing the browser never stops a search
- Run history reloaded from GitHub API on every page open (nothing lost)
- Per-search email notification address
- Auto-refresh while runs are active
"""

import time
import json
import zipfile
import io
import streamlit as st
import requests
import pandas as pd
from datetime import datetime, timezone

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="V Squared Sector Search Analysis",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Config ─────────────────────────────────────────────────────────────────────
GITHUB_TOKEN   = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO    = st.secrets.get("GITHUB_REPO", "mrdanielvlip-byte/companies-datascrape")
API_BASE       = "https://api.github.com"
HEADERS        = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}
ANTHROPIC_API_KEY = st.secrets.get("ANTHROPIC_API_KEY", "")
WORKFLOW_QUICK    = "pe_sourcing.yml"
WORKFLOW_DEEP     = "lift_maintenance_ocr.yml"
WORKFLOW_ESTIMATE = "estimate.yml"
DEFAULT_EMAIL     = "daniellipinski@mac.com"

# ── Session state ──────────────────────────────────────────────────────────────
if "pinned_runs" not in st.session_state:
    st.session_state.pinned_runs = []
if "active_tab" not in st.session_state:
    st.session_state.active_tab = 0
if "cached_runs" not in st.session_state:
    st.session_state.cached_runs = {}       # run_id → run dict
if "last_fetch" not in st.session_state:
    st.session_state.last_fetch = 0
if "run_inputs_store" not in st.session_state:
    st.session_state.run_inputs_store = {}  # run_id → inputs dict (sector, mode, region…)
if "pending_trigger" not in st.session_state:
    st.session_state.pending_trigger = None # inputs saved just before workflow dispatch
# Estimate state
if "estimate_triggered" not in st.session_state:
    st.session_state.estimate_triggered = False # True the moment Preview is clicked
if "estimate_run_id" not in st.session_state:
    st.session_state.estimate_run_id = None     # run_id once GitHub creates it
if "estimate_sector" not in st.session_state:
    st.session_state.estimate_sector = ""
if "estimate_result" not in st.session_state:
    st.session_state.estimate_result = None     # parsed estimate.json once complete
if "estimate_confirmed" not in st.session_state:
    st.session_state.estimate_confirmed = False # user pressed "Yes, run full search"
if "_estimate_max_companies" not in st.session_state:
    st.session_state["_estimate_max_companies"] = 0  # 0 = all companies
# Search universe state
if "_estimate_search_source" not in st.session_state:
    st.session_state["_estimate_search_source"] = "sic"
if "_estimate_reg_sources" not in st.session_state:
    st.session_state["_estimate_reg_sources"] = []
if "_estimate_reg_query" not in st.session_state:
    st.session_state["_estimate_reg_query"] = ""
# Company quality filter state
if "_estimate_min_age" not in st.session_state:
    st.session_state["_estimate_min_age"] = 0
if "_estimate_clean_charges" not in st.session_state:
    st.session_state["_estimate_clean_charges"] = False
if "_estimate_excluded_sics" not in st.session_state:
    st.session_state["_estimate_excluded_sics"] = []
# Trade body state — AUTO = let pipeline discover; list of keys = explicit selection
if "_estimate_trade_bodies" not in st.session_state:
    st.session_state["_estimate_trade_bodies"] = "AUTO"


# ── GitHub API helpers ─────────────────────────────────────────────────────────

def gh(path, method="GET", body=None):
    url = f"{API_BASE}{path}"
    try:
        if method == "POST":
            r = requests.post(url, headers=HEADERS, json=body, timeout=15)
        elif method == "PATCH":
            r = requests.patch(url, headers=HEADERS, json=body, timeout=15)
        else:
            r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 204:
            return {}
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"_error": str(e)}


def trigger_workflow(workflow_file, inputs):
    result = gh(f"/repos/{GITHUB_REPO}/actions/workflows/{workflow_file}/dispatches",
                method="POST", body={"ref": "main", "inputs": inputs})
    return "_error" not in result


def get_recent_runs(limit=20):
    """Fetch recent runs from both workflows, merged and sorted by date."""
    runs = []
    for wf in (WORKFLOW_QUICK, WORKFLOW_DEEP):
        data = gh(f"/repos/{GITHUB_REPO}/actions/workflows/{wf}/runs?per_page={limit}")
        for r in data.get("workflow_runs", []):
            r["_workflow_label"] = "Sector Search" if wf == WORKFLOW_QUICK else "Deep OCR"
            runs.append(r)
    runs.sort(key=lambda r: r["created_at"], reverse=True)
    return runs[:limit]


def get_run(run_id):
    cached = st.session_state.cached_runs.get(run_id)
    if cached and cached.get("status") == "completed":
        return cached   # completed runs don't change
    data = gh(f"/repos/{GITHUB_REPO}/actions/runs/{run_id}")
    if "_error" not in data:
        st.session_state.cached_runs[run_id] = data
    return data


def get_artifacts(run_id):
    data = gh(f"/repos/{GITHUB_REPO}/actions/runs/{run_id}/artifacts")
    return [a for a in data.get("artifacts", []) if not a.get("expired")]


def download_artifact(artifact_id):
    url = f"{API_BASE}/repos/{GITHUB_REPO}/actions/artifacts/{artifact_id}/zip"
    try:
        r = requests.get(url, headers=HEADERS, allow_redirects=True, timeout=60)
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def cancel_run(run_id):
    gh(f"/repos/{GITHUB_REPO}/actions/runs/{run_id}/cancel", method="POST")


def fetch_estimate_result(run_id) -> dict | None:
    """Download and parse estimate.json from a completed estimate workflow run."""
    import zipfile, io
    artifacts = get_artifacts(run_id)
    art = next((a for a in artifacts if a["name"] == "sector-estimate"), None)
    if not art:
        return None
    raw = download_artifact(art["id"])
    if not raw:
        return None
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            with z.open("estimate.json") as f:
                return json.loads(f.read())
    except Exception:
        return None


# ── Formatting helpers ─────────────────────────────────────────────────────────

def status_icon(status, conclusion):
    if status in ("queued", "in_progress"):
        return "⏳"
    if conclusion == "success":
        return "✅"
    if conclusion == "cancelled":
        return "🚫"
    if conclusion in ("failure", "timed_out"):
        return "❌"
    return "❓"


def fmt_duration(started, completed=None):
    try:
        start = datetime.strptime(started, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
        end   = (datetime.strptime(completed, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
                 if completed else datetime.now(timezone.utc))
        s = int((end - start).total_seconds())
        if s < 60:   return f"{s}s"
        if s < 3600: return f"{s//60}m {s%60}s"
        return f"{s//3600}h {(s%3600)//60}m"
    except Exception:
        return "—"


def run_display_name(run):
    name = run.get("display_title") or run.get("name") or f"Run {run.get('id', '?')}"
    return name[:50]


# ── Excel preview helper ──────────────────────────────────────────────────────

GRADE_COLOURS = {
    "Prime":             "#375623",   # dark green
    "High":              "#1F5C99",   # dark blue
    "Medium":            "#7B5B00",   # amber
    "Intelligence Only": "#555555",   # grey
}
GRADE_BG = {
    "Prime":             "#E2EFDA",
    "High":              "#DDEEFF",
    "Medium":            "#FFF2CC",
    "Intelligence Only": "#EEEEEE",
}

def parse_excel_preview(zip_bytes: bytes) -> dict | None:
    """
    Extract the PE Pipeline sheet from the downloaded zip artifact.
    Returns a dict with grade_counts, top_companies DataFrame, and summary stats.
    """
    try:
        from openpyxl import load_workbook
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            xlsx_names = [n for n in z.namelist() if n.endswith(".xlsx")]
            if not xlsx_names:
                return None
            xlsx_data = z.read(xlsx_names[0])

        wb = load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)

        if "PE Pipeline" not in wb.sheetnames:
            return None

        ws = wb["PE Pipeline"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 3 (index 2) is the header row
        if len(rows) < 4:
            return None

        headers = [str(h).strip() if h else "" for h in rows[2]]

        # Find key columns by header name
        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        ci_name   = col("Company Name")
        ci_grade  = col("Grade")
        ci_score  = col("Acq. Score")
        ci_rev    = col("Rev. Base £")
        ci_ebitda = col("EBITDA £")
        ci_epct   = col("EBITDA %")
        ci_sell   = col("SI Band")
        ci_family = col("Family")
        ci_pe     = col("PE Likelihood") or col("PE")
        ci_dirs   = col("Dirs")
        ci_age    = col("Age")
        ci_gscore = col("Growth Score")
        ci_perf   = col("Performance")
        ci_emp    = col("Employees")
        ci_fq     = col("Filing Quality")
        ci_maxage = col("Max Age")
        ci_corp   = col("Corp. Owner")
        ci_plike  = col("PE Likelihood")

        data_rows = rows[3:]   # skip title, subtitle, header
        records = []
        for r in data_rows:
            if not r or not any(r):
                continue
            name = r[ci_name] if ci_name is not None else ""
            if not name:
                continue
            records.append({
                "Company":     str(name),
                "Grade":       str(r[ci_grade]  or "") if ci_grade  is not None else "",
                "Score":       r[ci_score]  if ci_score  is not None else None,
                "Rev. Base":   str(r[ci_rev]    or "") if ci_rev    is not None else "",
                "EBITDA £":    str(r[ci_ebitda] or "") if ci_ebitda is not None else "",
                "EBITDA %":    str(r[ci_epct]   or "") if ci_epct   is not None else "",
                "Sell Intent": str(r[ci_sell]   or "") if ci_sell   is not None else "",
                "Family":      str(r[ci_family] or "") if ci_family is not None else "",
                "PE":          str(r[ci_pe]     or "") if ci_pe     is not None else "",
                "Dirs":        r[ci_dirs]  if ci_dirs  is not None else None,
                "Age (yr)":    r[ci_age]   if ci_age   is not None else None,
                "Growth":      r[ci_gscore] if ci_gscore is not None else None,
                "Performance": str(r[ci_perf] or "") if ci_perf is not None else "",
                "Employees":   r[ci_emp]   if ci_emp   is not None else None,
                "Filing":      str(r[ci_fq] or "") if ci_fq is not None else "",
                "Max Dir Age": r[ci_maxage] if ci_maxage is not None else None,
                "Corp. Owner": str(r[ci_corp] or "") if ci_corp is not None else "",
                "PE Likelihood": str(r[ci_plike] or "") if ci_plike is not None else "",
            })

        df = pd.DataFrame(records)
        if df.empty:
            return None

        # Grade distribution
        grade_counts = df["Grade"].value_counts().to_dict()
        grade_order  = ["Prime", "High", "Medium", "Intelligence Only"]
        grade_counts = {g: grade_counts.get(g, 0) for g in grade_order}

        # Sort all companies — Prime first, then High, sorted by score
        df["_score_num"] = pd.to_numeric(df["Score"], errors="coerce").fillna(0)
        grade_rank = {"Prime": 0, "High": 1, "Medium": 2, "Intelligence Only": 3}
        df["_grade_rank"] = df["Grade"].map(grade_rank).fillna(9)
        df_sorted = df.sort_values(["_grade_rank", "_score_num"], ascending=[True, False])

        top = df_sorted.head(25)
        display_cols = ["Company", "Grade", "Score", "Growth", "Performance",
                        "Rev. Base", "EBITDA £", "EBITDA %", "Sell Intent",
                        "Employees", "Family", "PE", "Dirs", "Age (yr)", "Filing",
                        "Max Dir Age", "Corp. Owner", "PE Likelihood"]
        # Only include columns that actually exist in the data
        display_cols = [c for c in display_cols if c in top.columns]
        top_df = top[display_cols].reset_index(drop=True)
        top_df.index = top_df.index + 1   # 1-based rank

        all_cols = [c for c in display_cols if c in df_sorted.columns]
        all_df = df_sorted[all_cols].reset_index(drop=True)
        all_df.index = all_df.index + 1

        return {"grade_counts": grade_counts, "top_df": top_df, "all_df": all_df, "total": len(df)}

    except Exception:
        return None


def _show_reenrich_panel(run_id: str, run: dict, inputs: dict):
    """
    Show a 'Re-run enrichment' expander on completed runs.
    Displays which modules were run originally, lets the user tick
    additional ones, and triggers a new --extras-only GitHub Actions run.
    """
    ALL_MODULES = [
        ("run_ocr",            "📄 Accounts OCR",      "Actual P&L from filed CH PDFs",               "~30 min"),
        ("run_contacts",       "📧 Director Contacts",  "Email inference + LinkedIn",                   "~20 min"),
        ("run_sell_signals",   "🚦 Sell Signals",       "Exit readiness, late filings, director churn", "~5 min"),
        ("run_contracts",      "🏛 Gov. Contracts",     "Contracts Finder + Find a Tender",             "~10 min"),
        ("run_digital",        "🌐 Digital Health",     "Website, domain age, LinkedIn, job postings",  "~20 min"),
        ("run_accreditations", "🔖 Accreditations",     "EA, CQC, FCA, ICO, SIA, ISO checks",          "~10 min"),
        ("run_competitor_map", "📍 Competitor Map",     "10 nearest rivals per company",                "~15 min"),
    ]

    sector = inputs.get("sector") or run.get("display_title", "")
    if not sector:
        return   # can't re-trigger without knowing the sector

    orig_modules = inputs.get("modules", {})   # dict from original trigger

    # Determine which modules were NOT run originally
    missing = [k for k, *_ in ALL_MODULES if not orig_modules.get(k, True)]

    with st.expander("🔄 Re-run / top-up enrichment modules", expanded=bool(missing)):
        st.markdown(
            "Select any modules to run (or re-run) against the existing search results. "
            "Discovery and financials are skipped — only the selected steps will execute."
        )

        notify_email = inputs.get("notify_email", DEFAULT_EMAIL)

        # Status badges for each module
        badge_cols = st.columns(len(ALL_MODULES))
        for i, (key, label, tip, est_time) in enumerate(ALL_MODULES):
            was_run = orig_modules.get(key, True)
            with badge_cols[i]:
                color = "#E2EFDA" if was_run else "#FFD6D6"
                text_color = "#1A5C2C" if was_run else "#7B0000"
                icon  = "✅" if was_run else "❌"
                st.markdown(
                    f"<div style='background:{color};border-radius:6px;padding:6px 8px;"
                    f"text-align:center;font-size:11px;color:{text_color};font-weight:600'>"
                    f"{icon} {label.split(' ',1)[1]}</div>",
                    unsafe_allow_html=True,
                )

        st.markdown("<br>", unsafe_allow_html=True)

        with st.form(key=f"reenrich_{run_id}"):
            st.markdown("**Which modules do you want to run?**")
            re_cols = st.columns(4)
            re_vals = {}
            for i, (key, label, tip, est_time) in enumerate(ALL_MODULES):
                was_run = orig_modules.get(key, True)
                # Pre-tick modules that were NOT run originally
                default = not was_run
                with re_cols[i % 4]:
                    re_vals[key] = st.checkbox(
                        label,
                        value=default,
                        help=f"{tip}  ·  Est. {est_time}",
                        key=f"re_{run_id}_{key}",
                    )

            re_email = st.text_input(
                "Notify email",
                value=notify_email,
                key=f"re_email_{run_id}",
            )

            selected_count = sum(re_vals.values())
            est_extra = sum(
                int(t.replace(" min","").replace("~",""))
                for (k, _, _, t) in ALL_MODULES if re_vals.get(k)
            )
            if selected_count:
                st.caption(f"{selected_count} module(s) selected · Est. ~{est_extra} min")

            submitted = st.form_submit_button(
                "🚀 Run selected modules",
                type="primary",
                disabled=(selected_count == 0),
            )

        if submitted and selected_count > 0:
            workflow_inputs = {
                "sector":      sector,
                "region":      inputs.get("region", ""),
                "min_revenue": inputs.get("min_revenue", ""),
                "notify_email": re_email,
                "extras_only": "true",   # skip steps 1–4
                **{k: "true" if v else "false" for k, v in re_vals.items()},
            }
            ok = trigger_workflow(WORKFLOW_QUICK, workflow_inputs)
            if ok:
                # Track which modules this new run includes
                combined = {k: (orig_modules.get(k, False) or re_vals.get(k, False)) for k, *_ in ALL_MODULES}
                st.session_state.pending_trigger = {
                    "sector": sector,
                    "region": inputs.get("region", ""),
                    "min_revenue": inputs.get("min_revenue", ""),
                    "notify_email": re_email,
                    "is_deep": False,
                    "modules": combined,
                    "triggered_at": datetime.now(timezone.utc).isoformat(),
                }
                st.success(f"✅ Re-enrichment run triggered for **{sector}**. You'll get an email at {re_email} when done.")
                time.sleep(3)
                st.cache_data.clear()
                st.rerun()
            else:
                st.error("Failed to trigger re-enrichment. Check your GitHub token in secrets.")


def _apply_post_filters(df: "pd.DataFrame", filters: dict) -> "pd.DataFrame":
    """Apply post-enrichment filters to a results DataFrame."""
    filtered = df.copy()

    # Director age
    min_dir = filters.get("min_dir_age", 0)
    max_dir = filters.get("max_dir_age", 999)
    if min_dir > 0 and "Max Dir Age" in filtered.columns:
        filtered["_max_dir_num"] = pd.to_numeric(filtered["Max Dir Age"], errors="coerce")
        filtered = filtered[filtered["_max_dir_num"].isna() | (filtered["_max_dir_num"] >= min_dir)]
    if max_dir < 999 and "Max Dir Age" in filtered.columns:
        if "_max_dir_num" not in filtered.columns:
            filtered["_max_dir_num"] = pd.to_numeric(filtered["Max Dir Age"], errors="coerce")
        filtered = filtered[filtered["_max_dir_num"].isna() | (filtered["_max_dir_num"] < max_dir)]

    # PE filter
    pe_rule = filters.get("pe", "Include all")
    if pe_rule != "Include all" and "PE Likelihood" in filtered.columns:
        if pe_rule == "Exclude High PE":
            filtered = filtered[filtered["PE Likelihood"] != "High"]
        elif pe_rule == "Exclude High + Medium":
            filtered = filtered[~filtered["PE Likelihood"].isin(["High", "Medium"])]
        elif pe_rule == "PE-backed only":
            filtered = filtered[filtered["PE Likelihood"].isin(["High", "Medium"])]

    # Family only
    if filters.get("family_only") and "Family" in filtered.columns:
        filtered = filtered[filtered["Family"].str.strip().isin(["✓", "Yes", "TRUE", "True", "1"])]

    # Min employees
    min_emp = filters.get("min_employees", 0)
    if min_emp > 0 and "Employees" in filtered.columns:
        filtered["_emp_num"] = pd.to_numeric(filtered["Employees"], errors="coerce")
        filtered = filtered[filtered["_emp_num"].isna() | (filtered["_emp_num"] >= min_emp)]

    # Min growth score
    min_gr = filters.get("min_growth", 0)
    if min_gr > 0 and "Growth" in filtered.columns:
        filtered["_gr_num"] = pd.to_numeric(filtered["Growth"], errors="coerce")
        filtered = filtered[filtered["_gr_num"].isna() | (filtered["_gr_num"] >= min_gr)]

    # Filing quality
    fq_list = filters.get("filing_quality", [])
    if fq_list and "Filing" in filtered.columns:
        filtered = filtered[filtered["Filing"].isin(fq_list) | (filtered["Filing"] == "")]

    # Min acquisition score
    min_acq = filters.get("min_acq_score", 0)
    if min_acq > 0 and "Score" in filtered.columns:
        filtered["_acq_num"] = pd.to_numeric(filtered["Score"], errors="coerce")
        filtered = filtered[filtered["_acq_num"].isna() | (filtered["_acq_num"] >= min_acq)]

    # Drop helper columns
    for c in ["_max_dir_num", "_emp_num", "_gr_num", "_acq_num"]:
        if c in filtered.columns:
            filtered = filtered.drop(columns=[c])

    return filtered


def show_excel_preview(zip_bytes: bytes, run_id: str):
    """Render grade distribution metrics + top companies table in Streamlit."""
    preview = parse_excel_preview(zip_bytes)
    if not preview:
        return

    st.divider()
    st.markdown("#### 📊 Results Preview")

    # ── Interactive filter panel ──────────────────────────────────────────────
    all_df = preview.get("all_df")  # full dataset (not just top 25)
    inputs = st.session_state.run_inputs_store.get(run_id, {})
    saved_filters = inputs.get("filters", {})

    with st.expander("🔍 Filter results", expanded=bool(any(v for k, v in saved_filters.items()
                                                             if k != "max_dir_age" and v not in (0, 999, "Include all", False, [], "")))):
        flt1, flt2, flt3, flt4 = st.columns(4)
        with flt1:
            _rv_min_dir = st.selectbox(
                "Min director age", ["Any", "45+", "50+", "55+", "60+", "65+"],
                index=0, key=f"rv_flt_dir_{run_id}",
            )
            _rv_dir_map = {"Any": 0, "45+": 45, "50+": 50, "55+": 55, "60+": 60, "65+": 65}
        with flt2:
            _rv_pe = st.selectbox(
                "PE ownership", ["Include all", "Exclude High PE", "Exclude High + Medium", "PE-backed only"],
                index=0, key=f"rv_flt_pe_{run_id}",
            )
        with flt3:
            _rv_min_emp = st.selectbox(
                "Min employees", ["Any", "5+", "10+", "20+", "50+", "100+"],
                index=0, key=f"rv_flt_emp_{run_id}",
            )
            _rv_emp_map = {"Any": 0, "5+": 5, "10+": 10, "20+": 20, "50+": 50, "100+": 100}
        with flt4:
            _rv_min_growth = st.selectbox(
                "Min Growth Score", ["Any", "25+", "40+", "55+", "75+"],
                index=0, key=f"rv_flt_gr_{run_id}",
            )
            _rv_gr_map = {"Any": 0, "25+": 25, "40+": 40, "55+": 55, "75+": 75}

        flt5, flt6, flt7, flt8 = st.columns(4)
        with flt5:
            _rv_min_acq = st.selectbox(
                "Min Acq. Score", ["Any", "30+", "50+", "65+", "80+"],
                index=0, key=f"rv_flt_acq_{run_id}",
            )
            _rv_acq_map = {"Any": 0, "30+": 30, "50+": 50, "65+": 65, "80+": 80}
        with flt6:
            _rv_family = st.checkbox("Family-owned only", value=False, key=f"rv_flt_fam_{run_id}")
        with flt7:
            _rv_perf = st.multiselect(
                "Performance",
                ["Strong Growth", "Growing", "Stable", "Declining", "Weak"],
                default=[], key=f"rv_flt_perf_{run_id}",
            )
        with flt8:
            _rv_filing = st.multiselect(
                "Filing quality",
                ["Full", "Medium", "Small-Full", "Small", "Exempt-Full", "Exempt-Small", "Abridged", "Micro"],
                default=[], key=f"rv_flt_fq_{run_id}",
            )

    # Build active filter dict from the interactive controls
    live_filters = {
        "min_dir_age":    _rv_dir_map[_rv_min_dir],
        "max_dir_age":    999,
        "pe":             _rv_pe,
        "family_only":    _rv_family,
        "min_employees":  _rv_emp_map[_rv_min_emp],
        "min_growth":     _rv_gr_map[_rv_min_growth],
        "filing_quality": _rv_filing,
        "min_acq_score":  _rv_acq_map[_rv_min_acq],
    }

    # Apply filters
    if all_df is not None and not all_df.empty:
        display_df = _apply_post_filters(all_df, live_filters)
        # Also filter by Performance multiselect
        if _rv_perf and "Performance" in display_df.columns:
            display_df = display_df[display_df["Performance"].isin(_rv_perf)]
    else:
        display_df = preview["top_df"]

    total_unfiltered = preview["total"]
    total_filtered   = len(display_df)

    # Grade distribution (from filtered data)
    gc_filtered = display_df["Grade"].value_counts().to_dict() if "Grade" in display_df.columns else {}
    grade_order = ["Prime", "High", "Medium", "Intelligence Only"]
    gc = {g: gc_filtered.get(g, 0) for g in grade_order}

    # Grade metric cards
    mcols = st.columns(4)
    for i, grade in enumerate(grade_order):
        count = gc.get(grade, 0)
        pct   = f"{count/total_filtered*100:.0f}%" if total_filtered > 0 else "0%"
        with mcols[i]:
            st.markdown(
                f"<div style='background:{GRADE_BG[grade]};border-radius:8px;padding:12px 16px;"
                f"text-align:center;border-left:4px solid {GRADE_COLOURS[grade]}'>"
                f"<div style='font-size:28px;font-weight:700;color:{GRADE_COLOURS[grade]}'>{count}</div>"
                f"<div style='font-size:12px;font-weight:600;color:{GRADE_COLOURS[grade]}'>{grade}</div>"
                f"<div style='font-size:11px;color:#888'>{pct} of {total_filtered}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    if total_filtered < total_unfiltered:
        st.markdown(
            f"**Showing {total_filtered:,} of {total_unfiltered:,} companies** "
            f"({total_unfiltered - total_filtered:,} filtered out)  ·  "
            f"Sorted by grade then acquisition score"
        )
    else:
        st.markdown(f"**All {total_filtered:,} companies** · Sorted by grade then acquisition score")

    # Sort for display
    grade_rank = {"Prime": 0, "High": 1, "Medium": 2, "Intelligence Only": 3}
    display_df["_score_num"] = pd.to_numeric(display_df["Score"], errors="coerce").fillna(0)
    display_df["_grade_rank"] = display_df["Grade"].map(grade_rank).fillna(9)
    display_df = display_df.sort_values(["_grade_rank", "_score_num"], ascending=[True, False])
    display_df = display_df.drop(columns=["_score_num", "_grade_rank"])
    display_df.index = range(1, len(display_df) + 1)

    # Choose display columns (hide internal ones)
    hide_cols = {"Max Dir Age", "Corp. Owner", "PE Likelihood"}
    display_cols = ["Company", "Grade", "Score", "Growth", "Performance",
                    "Rev. Base", "EBITDA £", "EBITDA %", "Sell Intent",
                    "Employees", "PE", "Family", "Dirs", "Age (yr)", "Filing"]
    display_cols = [c for c in display_cols if c in display_df.columns and c not in hide_cols]

    show_df = display_df[display_cols]

    # Colour-code Grade column
    def style_grade(val):
        bg = GRADE_BG.get(val, "#FFFFFF")
        fg = GRADE_COLOURS.get(val, "#000000")
        return f"background-color:{bg};color:{fg};font-weight:600"

    # Colour-code Performance column
    PERF_BG = {"Strong Growth": "#E2EFDA", "Growing": "#E2EFDA", "Stable": "#FFF2CC",
               "Declining": "#FFD6D6", "Weak": "#FFD6D6", "Insufficient Data": "#EEEEEE"}
    PERF_FG = {"Strong Growth": "#1A5C2C", "Growing": "#1A5C2C", "Stable": "#7B5B00",
               "Declining": "#7B0000", "Weak": "#7B0000", "Insufficient Data": "#888888"}

    def style_perf(val):
        bg = PERF_BG.get(str(val), "#FFFFFF")
        fg = PERF_FG.get(str(val), "#000000")
        return f"background-color:{bg};color:{fg};font-weight:600"

    style_subsets = [("Grade", style_grade)]
    if "Performance" in show_df.columns:
        style_subsets.append(("Performance", style_perf))

    styled = show_df.style
    for col_name, fn in style_subsets:
        styled = styled.applymap(fn, subset=[col_name])

    fmt = {"Score": lambda x: str(int(x)) if pd.notna(x) and x != "" else "-",
           "Dirs":  lambda x: str(int(x)) if pd.notna(x) and x != "" else "-",
           "Age (yr)": lambda x: f"{x:.0f}" if pd.notna(x) and x != "" else "-"}
    if "Growth" in show_df.columns:
        fmt["Growth"] = lambda x: str(int(x)) if pd.notna(x) and x != "" else "-"
    if "Employees" in show_df.columns:
        fmt["Employees"] = lambda x: str(int(x)) if pd.notna(x) and x != "" else "-"
    styled = styled.format(fmt)

    st.dataframe(styled, use_container_width=True, height=600)


# ── Load all recent runs (auto-pin any active ones) ───────────────────────────

@st.cache_data(ttl=30)
def load_all_runs():
    return get_recent_runs(20)


all_runs = load_all_runs()

# Match a pending trigger to the newest run we haven't catalogued yet
if st.session_state.pending_trigger:
    pt = st.session_state.pending_trigger
    triggered_at = pt.get("triggered_at", "")
    for r in all_runs[:5]:   # check the 5 most recent runs
        rid = r["id"]
        if rid not in st.session_state.run_inputs_store:
            # Only associate if run was created after (or within 10s before) trigger
            try:
                run_created = datetime.fromisoformat(r["created_at"].replace("Z", "+00:00"))
                trig_time   = datetime.fromisoformat(triggered_at.replace("Z", "+00:00"))
                if (run_created - trig_time).total_seconds() > -10:
                    st.session_state.run_inputs_store[rid] = pt
                    st.session_state.pending_trigger = None
                    break
            except Exception:
                pass

# Auto-pin any runs that are currently active (queued/in_progress)
for r in all_runs:
    if r["status"] in ("queued", "in_progress") and r["id"] not in st.session_state.pinned_runs:
        st.session_state.pinned_runs.append(r["id"])

# Build the run lookup dict
run_lookup = {r["id"]: r for r in all_runs}
for rid in st.session_state.pinned_runs:
    if rid not in run_lookup:
        fetched = get_run(rid)
        if "_error" not in fetched:
            run_lookup[rid] = fetched


# ── Header ─────────────────────────────────────────────────────────────────────

st.markdown("""
<h1 style='margin-bottom:0'>🔍 V Squared Sector Search Analysis</h1>
<p style='color:gray;margin-top:4px'>UK SME Acquisition Intelligence Platform &nbsp;·&nbsp; Searches run on GitHub — closing this browser never stops them</p>
""", unsafe_allow_html=True)

if not GITHUB_TOKEN:
    st.error("⚠️ GITHUB_TOKEN not configured in Streamlit secrets.")
    st.stop()

st.divider()


# ── Build tab list: New Search + one tab per pinned run ────────────────────────

pinned = st.session_state.pinned_runs  # ordered list of run_ids

def tab_name_for_run(rid):
    """Use stored sector name if available, else display_title (set by run-name: in YAML)."""
    inputs = st.session_state.run_inputs_store.get(rid, {})
    if inputs.get("sector"):
        return inputs["sector"]
    run = run_lookup.get(rid, {})
    return run_display_name(run) if run else f"Run {rid}"

tab_labels = ["➕ New Search"]
for rid in pinned:
    run  = run_lookup.get(rid, {})
    icon = status_icon(run.get("status", ""), run.get("conclusion"))
    name = tab_name_for_run(rid)
    tab_labels.append(f"{icon} {name[:28]}")

tabs = st.tabs(tab_labels)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 0 — New Search
# ══════════════════════════════════════════════════════════════════════════════

# ── AI Search Assistant — natural language → form config ──────────────────────
_AI_SYSTEM_PROMPT = """You are a PE deal-sourcing assistant. The user will describe what kind of companies they want to find.
Parse their description into a JSON config with these fields:

{
  "sector": "short sector description for SIC matching, e.g. 'fire safety systems'",
  "search_source": "sic" or "register" or "both",
  "reg_sources": [] or list of: "EA_WASTE", "EA_CARRIERS", "EA_ABSTRACTION", "EA_DISCHARGES", "CQC", "FCA",
  "reg_query": "" or keyword for register search,
  "region": "" or one of: "London", "South East", "South West", "East of England", "East Midlands", "West Midlands", "Yorkshire and The Humber", "North West", "North East", "Wales", "Scotland", "Northern Ireland",
  "min_revenue": "" or one of: "250000", "500000", "1000000", "2000000", "5000000",
  "min_age_years": 0 or integer (3, 5, 10, 15),
  "clean_charges_only": false or true,
  "min_dir_age": 0 or integer (45, 50, 55, 60, 65),
  "max_dir_age": 999 or integer (50, 55, 60, 65, 70, 75),
  "pe_filter": "Include all" or "Exclude High PE" or "Exclude High + Medium" or "PE-backed only",
  "family_only": false or true,
  "min_employees": 0 or integer (5, 10, 20, 50, 100),
  "min_growth_score": 0 or integer (25, 40, 55, 75),
  "min_acq_score": 0 or integer (30, 50, 65, 80),
  "run_deep": false or true,
  "explanation": "Brief explanation of how you interpreted their request"
}

Rules:
- If the user mentions care homes, healthcare → suggest CQC register
- If the user mentions waste, recycling, skip hire → suggest EA_WASTE register
- If the user mentions financial services, brokers, IFA → suggest FCA register
- If they want "established" or "mature" companies → set min_age_years to 10+
- If they mention succession, retirement, older directors → set min_dir_age to 55+
- If they say "no PE" or "independent" → set pe_filter to "Exclude High + Medium"
- If they want "growing" companies → set min_growth_score to 55
- If they mention a UK region, set it
- Default to search_source "sic" unless registers are clearly relevant
- Always set a sector even if using registers

Respond ONLY with the JSON object, no other text."""


def _ai_parse_search(user_input: str) -> dict | None:
    """Call Claude API to parse a natural language search description into form config."""
    if not ANTHROPIC_API_KEY:
        return None
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 1024,
                "system": _AI_SYSTEM_PROMPT,
                "messages": [{"role": "user", "content": user_input}],
            },
            timeout=30,
        )
        if resp.status_code == 200:
            content = resp.json()["content"][0]["text"]
            # Strip markdown code fences if present
            content = content.strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1] if "\n" in content else content[3:]
            if content.endswith("```"):
                content = content[:-3]
            return json.loads(content.strip())
    except Exception:
        pass
    return None


def _reset_estimate():
    st.session_state.estimate_triggered  = False
    st.session_state.estimate_run_id     = None
    st.session_state.estimate_result     = None
    st.session_state.estimate_sector     = ""
    st.session_state.estimate_confirmed  = False
    st.session_state["_estimate_max_companies"]  = 0
    st.session_state["_estimate_search_source"]  = "sic"
    st.session_state["_estimate_reg_sources"]    = []
    st.session_state["_estimate_reg_query"]      = ""
    st.session_state["_estimate_min_age"]        = 0
    st.session_state["_estimate_clean_charges"]  = False
    st.session_state["_estimate_excluded_sics"]  = []
    st.session_state["_estimate_trade_bodies"]   = "AUTO"

with tabs[0]:
    st.subheader("New Sector Search")
    st.caption("Each search runs independently on GitHub — you can start multiple and close this window at any time.")

    # ── AI Search Assistant ───────────────────────────────────────────────────
    if ANTHROPIC_API_KEY:
        with st.expander("💬 Describe your search in plain English", expanded=False):
            st.caption(
                "Tell me what you're looking for and I'll configure all the filters for you. "
                "For example: *\"Find established waste management companies in the North West, "
                "10+ years old, growing, no PE ownership, ideally family-run with older directors.\"*"
            )
            ai_col1, ai_col2 = st.columns([4, 1])
            with ai_col1:
                ai_input = st.text_area(
                    "What are you looking for?",
                    placeholder="e.g. 'I want to find care home businesses in the Midlands, at least 15 years old, with 20+ employees, growing revenue, and ideally family-owned with directors approaching retirement'",
                    height=80,
                    key="ai_search_input",
                    label_visibility="collapsed",
                )
            with ai_col2:
                ai_go = st.button("🤖 Configure", use_container_width=True, key="ai_go")

            if ai_go and ai_input.strip():
                with st.spinner("Parsing your search criteria…"):
                    ai_config = _ai_parse_search(ai_input.strip())
                if ai_config:
                    # Store parsed values into session state for form defaults
                    st.session_state["_ai_sector"]        = ai_config.get("sector", "")
                    st.session_state["_ai_search_source"] = ai_config.get("search_source", "sic")
                    st.session_state["_ai_reg_sources"]   = ai_config.get("reg_sources", [])
                    st.session_state["_ai_reg_query"]     = ai_config.get("reg_query", "")
                    st.session_state["_ai_region"]        = ai_config.get("region", "")
                    st.session_state["_ai_min_revenue"]   = ai_config.get("min_revenue", "")
                    st.session_state["_ai_min_age"]       = ai_config.get("min_age_years", 0)
                    st.session_state["_ai_clean_charges"] = ai_config.get("clean_charges_only", False)
                    st.session_state["_ai_min_dir_age"]   = ai_config.get("min_dir_age", 0)
                    st.session_state["_ai_max_dir_age"]   = ai_config.get("max_dir_age", 999)
                    st.session_state["_ai_pe_filter"]     = ai_config.get("pe_filter", "Include all")
                    st.session_state["_ai_family_only"]   = ai_config.get("family_only", False)
                    st.session_state["_ai_min_employees"] = ai_config.get("min_employees", 0)
                    st.session_state["_ai_min_growth"]    = ai_config.get("min_growth_score", 0)
                    st.session_state["_ai_min_acq"]       = ai_config.get("min_acq_score", 0)
                    st.session_state["_ai_run_deep"]      = ai_config.get("run_deep", False)

                    explanation = ai_config.get("explanation", "")
                    st.success(f"✅ **Search configured!** {explanation}")
                    st.caption("Review the form below — all fields have been pre-filled. Adjust anything, then click Preview.")
                    st.rerun()
                else:
                    st.error("Couldn't parse your description. Please try rephrasing or fill in the form manually.")

    # ── PHASE 1: Search form ───────────────────────────────────────────────────
    # Hide form as soon as Preview is clicked (estimate_triggered), not just
    # when we have a run_id — prevents the race condition where GitHub takes
    # longer than 4s to create the run and the form reappears.
    estimate_active = (
        st.session_state.estimate_triggered
        or st.session_state.estimate_run_id is not None
        or st.session_state.estimate_result is not None
        or st.session_state.estimate_confirmed
    )

    # ── Discovery-capable regulatory registers ────────────────────────────────
    _REG_OPTIONS = {
        "EA_WASTE":      ("🌿", "EA Waste Operations",    "Landfill, transfer, treatment, MRF, skip hire — waste mgmt firms with permits"),
        "EA_CARRIERS":   ("🚚", "EA Waste Carriers",      "Carriers, brokers and dealers of controlled waste — upper-tier registration"),
        "EA_ABSTRACTION":("💧", "EA Water Abstraction",   "Water abstraction licences — utilities, agriculture, industrial"),
        "EA_DISCHARGES": ("🏭", "EA Discharge Consents",  "Consents to discharge to watercourses — industrial, sewage, processing"),
        "CQC":           ("🏥", "CQC Providers",          "Regulated care homes, homecare, hospitals, dentists, GPs"),
        "FCA":           ("💼", "FCA Authorised Firms",   "Mortgage brokers, IFAs, insurers, consumer credit — FCA-regulated firms"),
    }

    if not estimate_active:
        with st.form("new_search", clear_on_submit=False):

            # ── Row 1: Sector + email ──────────────────────────────────────────
            _ai_sector = st.session_state.get("_ai_sector", "")
            col1, col2 = st.columns([2, 1])
            with col1:
                sector = st.text_input(
                    "Sector description *",
                    value=_ai_sector if _ai_sector else "",
                    placeholder='e.g. "fire safety systems", "HVAC contractors", "waste management"',
                    help="Free text — the pipeline auto-discovers SIC codes using curated maps and fuzzy matching.",
                )
            with col2:
                notify_email = st.text_input(
                    "Send results to",
                    value=DEFAULT_EMAIL,
                    help="Email address to notify when this specific search finishes.",
                )

            # ── Search universe ────────────────────────────────────────────────
            st.markdown("**Search universe** — where to look for companies")
            src_col, reg_col = st.columns([1, 2])
            with src_col:
                _ai_src = st.session_state.get("_ai_search_source", "sic")
                _src_idx = {"sic": 0, "register": 1, "both": 2}.get(_ai_src, 0)
                search_source = st.radio(
                    "Discover from",
                    options=["sic", "register", "both"],
                    format_func=lambda x: {
                        "sic":      "📋  SIC codes  (all matching companies)",
                        "register": "🏛  Regulatory register  (registered firms only)",
                        "both":     "📋 + 🏛  SIC + register  (broadest coverage)",
                    }[x],
                    index=_src_idx,
                    help=(
                        "**SIC codes** — broadest. Finds every active UK company in the sector.\n\n"
                        "**Regulatory register** — highest quality signal. Only firms registered "
                        "with the relevant authority (EA, CQC, FCA, etc.).\n\n"
                        "**Both** — runs SIC search first then merges in register results, "
                        "deduplicating by company number."
                    ),
                )
            with reg_col:
                if search_source in ("register", "both"):
                    st.caption("Select the register(s) relevant to this sector:")
                    reg_cols = st.columns(3)
                    reg_sources_sel = []
                    for i, (key, (icon, name, desc)) in enumerate(_REG_OPTIONS.items()):
                        with reg_cols[i % 3]:
                            if st.checkbox(
                                f"{icon} {name}",
                                value=False,
                                help=desc,
                                key=f"reg_{key}",
                            ):
                                reg_sources_sel.append(key)
                    reg_query = st.text_input(
                        "Register keyword (optional)",
                        placeholder='Leave blank for all entries, or enter e.g. "drainage"',
                        help="Filter the register by keyword before cross-referencing with Companies House.",
                    )
                else:
                    reg_sources_sel = []
                    reg_query       = ""

            # ── Trade body member discovery ────────────────────────────────────
            st.markdown("**Trade & industry body members**")
            st.caption(
                "The pipeline automatically searches for relevant UK trade associations "
                "when your sector is run. Known bodies are pre-loaded below — tick to "
                "confirm or leave on AUTO to let the pipeline search dynamically."
            )

            # Instant suggestions from known bodies based on what's typed so far
            _known_suggestions = []
            if sector.strip():
                try:
                    from trade_body_finder import suggest_for_streamlit
                    _known_suggestions = suggest_for_streamlit(sector.strip())
                except ImportError:
                    pass

            tb_col1, tb_col2 = st.columns([3, 1])
            with tb_col1:
                if _known_suggestions:
                    st.caption("✅ **Known bodies matched to this sector:**")
                    _tb_sel = []
                    tb_grid = st.columns(min(3, len(_known_suggestions)))
                    for idx, body in enumerate(_known_suggestions):
                        with tb_grid[idx % 3]:
                            cnt_str   = (f" ~{body['member_count_est']} members"
                                         if body.get("member_count_est") else "")
                            disc_icon = "✅" if body["discoverable"] else "🚫"
                            checked   = st.checkbox(
                                f"{disc_icon} {body['name']}{cnt_str}",
                                value=body["discoverable"],
                                help=(
                                    f"{body['name']}\n{body['url']}"
                                    + (f"\n\n⚠️ {body['note']}" if body.get("note") else "")
                                ),
                                key=f"tb_{body['key']}",
                                disabled=not body["discoverable"],
                            )
                            if checked and body["discoverable"]:
                                _tb_sel.append(body["key"])
                    _tb_sel_final = _tb_sel if _tb_sel else ["AUTO"]
                else:
                    st.caption(
                        "No pre-loaded bodies for this sector — "
                        "pipeline will search automatically when the run starts."
                    )
                    _tb_sel_final = ["AUTO"]
            with tb_col2:
                st.caption("🔍 Auto-search always also runs for any bodies not listed above.")

            # ── Advanced filters ──────────────────────────────────────────────
            with st.expander("Advanced filters (optional)"):
                _ai_region = st.session_state.get("_ai_region", "")
                _ai_min_rev = st.session_state.get("_ai_min_revenue", "")
                _region_opts = [
                    "", "London", "South East", "South West", "East of England",
                    "East Midlands", "West Midlands", "Yorkshire and The Humber",
                    "North West", "North East", "Wales", "Scotland", "Northern Ireland",
                ]
                _rev_opts = ["", "250000", "500000", "1000000", "2000000", "5000000"]
                fc1, fc2 = st.columns(2)
                with fc1:
                    region = st.selectbox("UK region", _region_opts,
                        index=_region_opts.index(_ai_region) if _ai_region in _region_opts else 0,
                        help="Leave blank for all UK.")
                with fc2:
                    min_revenue = st.selectbox("Minimum estimated revenue", _rev_opts,
                        format_func=lambda x: "No minimum" if x == "" else f"£{int(x):,}",
                        index=_rev_opts.index(str(_ai_min_rev)) if str(_ai_min_rev) in _rev_opts else 0)

                st.markdown("**Company quality filters**")
                qf1, qf2, qf3 = st.columns(3)
                with qf1:
                    _MIN_AGE_MAP = {"Any age": 0, "3+ years": 3, "5+ years": 5, "10+ years": 10, "15+ years": 15}
                    _ai_min_age = st.session_state.get("_ai_min_age", 0)
                    _age_opts = list(_MIN_AGE_MAP.keys())
                    _age_rev  = {v: k for k, v in _MIN_AGE_MAP.items()}
                    _age_idx  = _age_opts.index(_age_rev.get(_ai_min_age, "Any age"))
                    min_age_sel = st.selectbox(
                        "Minimum company age",
                        options=_age_opts,
                        index=_age_idx,
                        help="Exclude recently incorporated companies. Useful for sectors where new entrants are less likely PE targets.",
                    )
                    min_age_yrs = _MIN_AGE_MAP[min_age_sel]
                with qf2:
                    clean_charges = st.checkbox(
                        "Clean charges register only",
                        value=False,
                        help="Only include companies with zero outstanding charges on the CH register — "
                             "a strong dealability signal.",
                    )
                with qf3:
                    st.caption(
                        "Further SIC refinement available after preview — "
                        "untick individual SIC codes before confirming the full run."
                    )

                st.markdown("**Director & ownership filters** *(applied to results after enrichment)*")
                df1, df2, df3, df4 = st.columns(4)
                with df1:
                    _DIR_AGE_MAP = {"Any": 0, "45+": 45, "50+": 50, "55+": 55, "60+": 60, "65+": 65}
                    _ai_mda = st.session_state.get("_ai_min_dir_age", 0)
                    _da_rev = {v: k for k, v in _DIR_AGE_MAP.items()}
                    _da_idx = list(_DIR_AGE_MAP.keys()).index(_da_rev.get(_ai_mda, "Any"))
                    min_dir_age_sel = st.selectbox(
                        "Min director age",
                        options=list(_DIR_AGE_MAP.keys()),
                        index=_da_idx,
                        help="Filter for companies where the oldest director is at least this age — "
                             "higher director age = stronger succession signal.",
                        key="flt_min_dir_age",
                    )
                    min_dir_age = _DIR_AGE_MAP[min_dir_age_sel]
                with df2:
                    _DIR_MAX_MAP = {"Any": 999, "Under 50": 50, "Under 55": 55, "Under 60": 60,
                                    "Under 65": 65, "Under 70": 70, "Under 75": 75}
                    _ai_maxda = st.session_state.get("_ai_max_dir_age", 999)
                    _dmax_rev = {v: k for k, v in _DIR_MAX_MAP.items()}
                    _dmax_idx = list(_DIR_MAX_MAP.keys()).index(_dmax_rev.get(_ai_maxda, "Any"))
                    max_dir_age_sel = st.selectbox(
                        "Max director age",
                        options=list(_DIR_MAX_MAP.keys()),
                        index=_dmax_idx,
                        help="Exclude companies where the oldest director exceeds this age.",
                        key="flt_max_dir_age",
                    )
                    max_dir_age = _DIR_MAX_MAP[max_dir_age_sel]
                with df3:
                    _pe_opts = ["Include all", "Exclude High PE", "Exclude High + Medium", "PE-backed only"]
                    _ai_pe = st.session_state.get("_ai_pe_filter", "Include all")
                    _pe_idx = _pe_opts.index(_ai_pe) if _ai_pe in _pe_opts else 0
                    pe_filter_sel = st.selectbox(
                        "PE ownership",
                        options=_pe_opts,
                        index=_pe_idx,
                        help="Filter companies by PE ownership likelihood. "
                             "'Exclude High PE' removes companies with strong PE signals.",
                        key="flt_pe",
                    )
                with df4:
                    _ai_fam = st.session_state.get("_ai_family_only", False)
                    family_filter = st.checkbox(
                        "Family-owned only",
                        value=_ai_fam,
                        help="Only show companies flagged as likely family-owned "
                             "(shared surnames, long tenures, small boards).",
                        key="flt_family",
                    )

                st.markdown("**Performance & size filters** *(applied to results after enrichment)*")
                pf1, pf2, pf3, pf4 = st.columns(4)
                with pf1:
                    _EMP_MAP = {"Any": 0, "5+": 5, "10+": 10, "20+": 20, "50+": 50, "100+": 100}
                    _ai_emp = st.session_state.get("_ai_min_employees", 0)
                    _emp_rev = {v: k for k, v in _EMP_MAP.items()}
                    _emp_idx = list(_EMP_MAP.keys()).index(_emp_rev.get(_ai_emp, "Any"))
                    min_employees_sel = st.selectbox(
                        "Min employees",
                        options=list(_EMP_MAP.keys()),
                        index=_emp_idx,
                        help="Filter by minimum employee count (where known).",
                        key="flt_min_emp",
                    )
                    min_employees = _EMP_MAP[min_employees_sel]
                with pf2:
                    _GROWTH_MAP = {"Any": 0, "25+ (not weak)": 25, "40+ (stable+)": 40,
                                   "55+ (growing+)": 55, "75+ (strong)": 75}
                    _ai_gr = st.session_state.get("_ai_min_growth", 0)
                    _gr_rev = {v: k for k, v in _GROWTH_MAP.items()}
                    _gr_idx = list(_GROWTH_MAP.keys()).index(_gr_rev.get(_ai_gr, "Any"))
                    min_growth_sel = st.selectbox(
                        "Min Growth Score",
                        options=list(_GROWTH_MAP.keys()),
                        index=_gr_idx,
                        help="Filter by composite Growth Score (0–100). "
                             "Higher = stronger growth trajectory across all available signals.",
                        key="flt_min_growth",
                    )
                    min_growth = _GROWTH_MAP[min_growth_sel]
                with pf3:
                    filing_filter_sel = st.multiselect(
                        "Filing quality",
                        options=["Full", "Medium", "Small-Full", "Small", "Exempt-Full",
                                 "Exempt-Small", "Abridged", "Micro", "Dormant"],
                        default=[],
                        help="Only show companies with these filing types. "
                             "Leave empty to include all. Full/Medium have the most data.",
                        key="flt_filing",
                    )
                with pf4:
                    _ACQ_MAP = {"Any": 0, "30+": 30, "50+ (Medium+)": 50, "65+ (High+)": 65, "80+ (Prime)": 80}
                    _ai_acq = st.session_state.get("_ai_min_acq", 0)
                    _acq_rev = {v: k for k, v in _ACQ_MAP.items()}
                    _acq_idx = list(_ACQ_MAP.keys()).index(_acq_rev.get(_ai_acq, "Any"))
                    min_acq_score_sel = st.selectbox(
                        "Min Acquisition Score",
                        options=list(_ACQ_MAP.keys()),
                        index=_acq_idx,
                        help="Filter by overall Acquisition Score grade.",
                        key="flt_min_acq",
                    )
                    min_acq_score = _ACQ_MAP[min_acq_score_sel]

            # ── Enrichment modules ────────────────────────────────────────────
            st.markdown("**Enrichment modules** — tick what to include in this run")
            st.caption("Core pipeline (search → filter → enrich → financials) always runs. Untick modules to speed up.")

            MODULES = [
                ("run_ocr",            "📄 Accounts OCR",      "Actual P&L figures from filed CH PDFs",               "~30 min", True),
                ("run_contacts",       "📧 Director Contacts",  "Email inference + LinkedIn for each director",         "~20 min", True),
                ("run_sell_signals",   "🚦 Sell Signals",       "Exit readiness: late filings, director churn, tenure", "~5 min",  True),
                ("run_contracts",      "🏛 Gov. Contracts",     "Contracts Finder + Find a Tender lookups",             "~10 min", True),
                ("run_digital",        "🌐 Digital Health",     "Website, domain age, LinkedIn, job postings",          "~20 min", True),
                ("run_accreditations", "🔖 Accreditations",     "EA, CQC, FCA, ICO, SIA, ISO register checks",         "~10 min", True),
                ("run_competitor_map", "📍 Competitor Map",     "10 nearest rivals per company (geographic + SIC)",     "~15 min", True),
            ]

            module_vals = {}
            mod_cols = st.columns(4)
            for i, (key, label, tip, est_time, default) in enumerate(MODULES):
                with mod_cols[i % 4]:
                    module_vals[key] = st.checkbox(
                        f"{label}",
                        value=default,
                        help=f"{tip}  ·  Est. {est_time}",
                        key=f"mod_{key}",
                    )

            checked_count = sum(module_vals.values())
            est_total = 10 + sum(
                int(t.replace(" min","").replace("~",""))
                for (k, _, _, t, _) in MODULES if module_vals.get(k)
            )
            st.caption(
                f"**{checked_count}/{len(MODULES)} modules selected** "
                f"· Estimated run time: ~{est_total} min"
                + (" · All Excel tabs will be populated" if checked_count == len(MODULES) else
                   " · Some Excel tabs will be blank (run Re-Enrich later to fill them in)")
            )

            st.divider()
            deep_col, preview_col, _ = st.columns([1, 1, 1])
            with deep_col:
                run_deep = st.checkbox("Deep OCR run (~5.5 hrs, most thorough)")
            with preview_col:
                preview_clicked = st.form_submit_button(
                    "🔍 Preview Companies", use_container_width=True,
                    help="Run a quick ~2 min estimate to see company count and SIC accuracy before committing.",
                )

        if preview_clicked:
            if not sector.strip():
                st.error("Please enter a sector description.")
            elif search_source in ("register", "both") and not reg_sources_sel:
                st.error("Please select at least one regulatory register, or switch to 'SIC codes' mode.")
            else:
                st.session_state.estimate_triggered  = True
                st.session_state.estimate_sector     = sector.strip()
                st.session_state._estimate_email     = notify_email
                st.session_state._estimate_region    = region
                st.session_state._estimate_rev       = min_revenue
                st.session_state._estimate_deep      = run_deep
                st.session_state._estimate_modules   = dict(module_vals)
                st.session_state["_estimate_search_source"] = search_source
                st.session_state["_estimate_reg_sources"]   = reg_sources_sel
                st.session_state["_estimate_reg_query"]     = reg_query
                st.session_state["_estimate_min_age"]       = min_age_yrs
                st.session_state["_estimate_clean_charges"] = clean_charges
                st.session_state["_estimate_trade_bodies"]  = _tb_sel_final
                # Post-enrichment filters (applied to results, not workflow inputs)
                st.session_state["_filter_min_dir_age"]    = min_dir_age
                st.session_state["_filter_max_dir_age"]    = max_dir_age
                st.session_state["_filter_pe"]             = pe_filter_sel
                st.session_state["_filter_family_only"]    = family_filter
                st.session_state["_filter_min_employees"]  = min_employees
                st.session_state["_filter_min_growth"]     = min_growth
                st.session_state["_filter_filing_quality"] = filing_filter_sel
                st.session_state["_filter_min_acq_score"]  = min_acq_score
                ok = trigger_workflow(WORKFLOW_ESTIMATE, {"sector": sector.strip()})
                if not ok:
                    _reset_estimate()
                    st.error("Failed to trigger estimate. Check your GitHub token in secrets.")
                else:
                    st.rerun()

    # ── PHASE 2: Estimate in progress (triggered but may not have run_id yet) ──
    elif (st.session_state.estimate_triggered or st.session_state.estimate_run_id) \
            and st.session_state.estimate_result is None \
            and not st.session_state.estimate_confirmed:

        st.info(f"🔍 Estimating company universe for **{st.session_state.estimate_sector}** …")

        # If we don't have a run_id yet, poll GitHub until the run appears
        if st.session_state.estimate_run_id is None:
            wf_runs = gh(f"/repos/{GITHUB_REPO}/actions/workflows/{WORKFLOW_ESTIMATE}/runs?per_page=1")
            runs = wf_runs.get("workflow_runs", [])
            if runs:
                st.session_state.estimate_run_id = runs[0]["id"]
                st.session_state.estimate_triggered = True  # keep active
            else:
                st.progress(0.02, text="Waiting for GitHub to start the estimate job…")
                st.caption("Auto-checks every 5 s")
                if st.button("✕ Cancel", key="cancel_est_wait"):
                    _reset_estimate()
                    st.rerun()
                time.sleep(5)
                st.rerun()
        else:
            run_id = st.session_state.estimate_run_id
            run    = get_run(run_id)
            status = run.get("status", "unknown")
            conc   = run.get("conclusion")

            if status in ("queued", "in_progress"):
                started_str = run.get("run_started_at") or run.get("created_at", "")
                try:
                    started_dt  = datetime.fromisoformat(started_str.replace("Z", "+00:00"))
                    elapsed_sec = (datetime.now(timezone.utc) - started_dt).total_seconds()
                    pct     = min(elapsed_sec / 120, 0.95)
                    remain  = max(0, int((120 - elapsed_sec) // 60))
                    bar_txt = f"~{int(elapsed_sec//60)}m {int(elapsed_sec%60)}s elapsed · ~{remain} min remaining"
                except Exception:
                    pct, bar_txt = 0.05, "Starting up…"

                st.progress(pct, text=bar_txt)
                st.caption("Auto-refreshes every 10 s")
                if st.button("✕ Cancel estimate", key="cancel_est"):
                    cancel_run(run_id)
                    _reset_estimate()
                    st.rerun()
                time.sleep(10)
                st.rerun()

            elif status == "completed" and conc == "success":
                result = fetch_estimate_result(run_id)
                if result:
                    st.session_state.estimate_result = result
                    st.rerun()
                else:
                    st.error("Estimate finished but could not read results. Try again.")
                    if st.button("Start over", key="est_retry"):
                        _reset_estimate()
                        st.rerun()
            else:
                st.error(f"Estimate ended with: **{conc or status}**. Try again or use a different sector description.")
                if st.button("Start over", key="est_fail"):
                    _reset_estimate()
                    st.rerun()

    # ── PHASE 3: Show estimate results and ask to confirm ─────────────────────
    elif st.session_state.estimate_result is not None \
            and not st.session_state.estimate_confirmed:
        r          = st.session_state.estimate_result
        sector     = r.get("sector", st.session_state.estimate_sector)
        total      = r.get("total_companies", 0)
        acc_pct    = r.get("accuracy_pct", 0)
        acc_lbl    = r.get("accuracy_label", "")
        source     = r.get("match_source", "fuzzy")
        sic_bkd    = r.get("sic_breakdown", [])
        samples    = r.get("sample_companies", [])
        api_errors = r.get("api_errors", 0)

        # Accuracy colour
        if acc_pct >= 80:
            acc_colour = "🟢"
        elif acc_pct >= 60:
            acc_colour = "🟡"
        else:
            acc_colour = "🔴"

        st.success(f"**Estimate complete for: {sector}**")

        if api_errors > 0:
            st.warning(
                f"⚠️ {api_errors} SIC code(s) couldn't be counted due to API rate limits — "
                "the total shown may be **understated**. The full search will not be affected."
            )

        # ── Top metrics ────────────────────────────────────────────────────────
        m1, m2, m3 = st.columns(3)
        m1.metric("Estimated companies", f"{total:,}")
        m2.metric("Sector accuracy", f"{acc_pct}%")
        m3.metric("SIC match method", "Curated map" if source == "curated" else "Fuzzy match")

        st.caption(f"{acc_colour} {acc_lbl}")

        # ── Active post-enrichment filters summary ────────────────────────────
        _active_filters = []
        _f_min_dir = st.session_state.get("_filter_min_dir_age", 0)
        _f_max_dir = st.session_state.get("_filter_max_dir_age", 999)
        _f_pe      = st.session_state.get("_filter_pe", "Include all")
        _f_family  = st.session_state.get("_filter_family_only", False)
        _f_min_emp = st.session_state.get("_filter_min_employees", 0)
        _f_min_gr  = st.session_state.get("_filter_min_growth", 0)
        _f_filing  = st.session_state.get("_filter_filing_quality", [])
        _f_min_acq = st.session_state.get("_filter_min_acq_score", 0)

        if _f_min_dir > 0:
            _active_filters.append(f"Director age ≥ {_f_min_dir}")
        if _f_max_dir < 999:
            _active_filters.append(f"Director age < {_f_max_dir}")
        if _f_pe != "Include all":
            _active_filters.append(f"PE: {_f_pe}")
        if _f_family:
            _active_filters.append("Family-owned only")
        if _f_min_emp > 0:
            _active_filters.append(f"Employees ≥ {_f_min_emp}")
        if _f_min_gr > 0:
            _active_filters.append(f"Growth Score ≥ {_f_min_gr}")
        if _f_filing:
            _active_filters.append(f"Filing: {', '.join(_f_filing)}")
        if _f_min_acq > 0:
            _active_filters.append(f"Acq. Score ≥ {_f_min_acq}")

        if _active_filters:
            st.info(
                f"**Post-enrichment filters active** ({len(_active_filters)}): "
                + "  ·  ".join(_active_filters)
                + "\n\nThese will be applied to results after the pipeline completes. "
                f"The pipeline will process all ~{total:,} companies, then filter the Excel preview."
            )

        st.divider()

        # ── SIC code breakdown + interactive exclusion form ───────────────────
        # Everything interactive (SIC exclusion + confirm) goes in one form so
        # all checkbox values are captured atomically on submission.
        with st.form("phase3_confirm"):
            if sic_bkd:
                st.markdown(
                    "**Matched SIC codes** — untick any to exclude from the search"
                )
                st.caption(
                    "Each SIC code was identified as part of this sector. "
                    "Untick codes that don't belong — e.g. if a broad SIC is pulling in "
                    "unrelated businesses."
                )
                sic_cols = st.columns(2)
                for i, s in enumerate(sic_bkd):
                    pct_of_total = (s["count"] / total * 100) if total else 0
                    with sic_cols[i % 2]:
                        st.checkbox(
                            f"`{s['code']}` {s['description']}  —  "
                            f"**{s['count']:,}** cos ({pct_of_total:.0f}%)",
                            value=True,
                            key=f"sic_keep_{s['code']}",
                        )
            else:
                st.info("No SIC breakdown available for this search.")

            if samples:
                st.markdown("**Sample companies found**")
                st.caption("  ·  ".join(samples[:10]))

            st.divider()
            st.markdown("**Confirm and customise your search**")

            # ── Company count limiter ──────────────────────────────────────────
            _COUNT_OPTIONS = {
                "All companies":          0,
                "Top 25  (quick test)":  25,
                "Top 50  (quick test)":  50,
                "Top 100":              100,
                "Top 250":              250,
                "Top 500":              500,
            }
            if total > 500:
                _default_lbl = "Top 250"
            elif total > 100:
                _default_lbl = "Top 100"
            else:
                _default_lbl = "All companies"

            _selected_lbl = st.selectbox(
                "How many companies to process?",
                options=list(_COUNT_OPTIONS.keys()),
                index=list(_COUNT_OPTIONS.keys()).index(_default_lbl),
                help=(
                    "Limit the number of companies enriched in this run. "
                    "Choose a smaller number for a quick test — you can always "
                    "top up the rest afterwards using the Re-Enrich panel."
                ),
            )
            _max_n = _COUNT_OPTIONS[_selected_lbl]

            if _max_n and _max_n < total:
                st.caption(
                    f"ℹ️ Will process the first **{_max_n:,}** of ~{total:,} estimated "
                    f"companies (≈ {_max_n / total * 100:.0f}% of the dataset)."
                )
            else:
                st.caption(f"ℹ️ Will process all ~{total:,} estimated companies.")

            # ── Submit buttons ─────────────────────────────────────────────────
            yes_col, no_col = st.columns([1, 1])
            with yes_col:
                yes_clicked = st.form_submit_button(
                    "✅ Yes — Run Full Search", type="primary", use_container_width=True,
                )
            with no_col:
                no_clicked = st.form_submit_button(
                    "✏️ No — Change Sector", use_container_width=True,
                )

        if yes_clicked:
            # Capture excluded SICs from form state
            _excl = [
                s["code"] for s in sic_bkd
                if not st.session_state.get(f"sic_keep_{s['code']}", True)
            ]
            st.session_state["_estimate_excluded_sics"]  = _excl
            st.session_state["_estimate_max_companies"]  = _max_n
            st.session_state.estimate_confirmed = True
            st.rerun()
        if no_clicked:
            _reset_estimate()
            st.rerun()

    # ── PHASE 4: Confirmed — trigger the real pipeline ─────────────────────────
    elif st.session_state.estimate_confirmed:
        sector        = st.session_state.estimate_sector
        notify_email  = st.session_state.get("_estimate_email", DEFAULT_EMAIL)
        region        = st.session_state.get("_estimate_region", "")
        min_revenue   = st.session_state.get("_estimate_rev", "")
        run_deep      = st.session_state.get("_estimate_deep", False)
        max_companies = st.session_state.get("_estimate_max_companies", 0)
        search_source = st.session_state.get("_estimate_search_source", "sic")
        reg_sources   = st.session_state.get("_estimate_reg_sources", [])
        reg_query     = st.session_state.get("_estimate_reg_query", "")
        min_age       = st.session_state.get("_estimate_min_age", 0)
        clean_charges = st.session_state.get("_estimate_clean_charges", False)
        excluded_sics = st.session_state.get("_estimate_excluded_sics", [])
        trade_bodies  = st.session_state.get("_estimate_trade_bodies", "AUTO")
        modules       = st.session_state.get("_estimate_modules", {k: True for k, *_ in [
            ("run_ocr",), ("run_contacts",), ("run_sell_signals",),
            ("run_contracts",), ("run_digital",), ("run_accreditations",), ("run_competitor_map",),
        ]})

        with st.spinner(f"Launching full pipeline for '{sector}'…"):
            if run_deep:
                ok = trigger_workflow(WORKFLOW_DEEP, {
                    "sector": sector, "notify_email": notify_email,
                })
                label = f"{sector} (Deep OCR)"
                st.session_state.pending_trigger = {
                    "sector": sector, "region": "", "min_revenue": "",
                    "notify_email": notify_email, "is_deep": True,
                    "modules": {k: True for k in modules},
                    "triggered_at": datetime.now(timezone.utc).isoformat(),
                }
            else:
                workflow_inputs = {
                    "sector":             sector,
                    "region":             region,
                    "min_revenue":        min_revenue,
                    "notify_email":       notify_email,
                    "extras_only":        "false",
                    "max_companies":      str(max_companies) if max_companies else "",
                    "search_source":      search_source,
                    "reg_sources":        ",".join(reg_sources) if reg_sources else "",
                    "reg_query":          reg_query or "",
                    "min_age_years":      str(min_age) if min_age else "",
                    "clean_charges_only": "true" if clean_charges else "false",
                    "excluded_sics":      ",".join(excluded_sics) if excluded_sics else "",
                    "trade_bodies":       (
                        ",".join(trade_bodies)
                        if isinstance(trade_bodies, list) else str(trade_bodies or "AUTO")
                    ),
                    **{k: "true" if v else "false" for k, v in modules.items()},
                }
                ok = trigger_workflow(WORKFLOW_QUICK, workflow_inputs)
                label = sector
                st.session_state.pending_trigger = {
                    "sector": sector, "region": region, "min_revenue": min_revenue,
                    "notify_email": notify_email, "is_deep": False,
                    "modules": modules,
                    "search_source": search_source,
                    "reg_sources": reg_sources,
                    "triggered_at": datetime.now(timezone.utc).isoformat(),
                    "filters": {
                        "min_dir_age":    st.session_state.get("_filter_min_dir_age", 0),
                        "max_dir_age":    st.session_state.get("_filter_max_dir_age", 999),
                        "pe":             st.session_state.get("_filter_pe", "Include all"),
                        "family_only":    st.session_state.get("_filter_family_only", False),
                        "min_employees":  st.session_state.get("_filter_min_employees", 0),
                        "min_growth":     st.session_state.get("_filter_min_growth", 0),
                        "filing_quality": st.session_state.get("_filter_filing_quality", []),
                        "min_acq_score":  st.session_state.get("_filter_min_acq_score", 0),
                    },
                }

        if ok:
            st.success(f"✅ Full pipeline triggered for **{label}**.")
            st.info(f"📧 You'll get an email at **{notify_email}** when it's done.")
            _reset_estimate()
            time.sleep(4)
            st.cache_data.clear()
            st.rerun()
        else:
            st.error("Failed to trigger pipeline. Check your GitHub token in secrets.")
            st.session_state.estimate_confirmed = False

    # ── Recent runs summary table ──────────────────────────────────────────────
    st.divider()
    st.subheader("All Runs")
    st.caption("Click **Pin as tab** to open a run in its own tab for detailed status and download.")

    if not all_runs:
        st.info("No runs yet — trigger one above.")
    else:
        for run in all_runs[:15]:
            rid     = run["id"]
            status  = run["status"]
            conc    = run.get("conclusion")
            icon    = status_icon(status, conc)
            name    = run_display_name(run)
            label   = run.get("_workflow_label", "")
            date    = run["created_at"][:10]
            dur     = fmt_duration(run["created_at"], run.get("updated_at") if status == "completed" else None)

            c1, c2, c3 = st.columns([5, 1, 1])
            with c1:
                st.markdown(f"**{icon} {name}** &nbsp; `{label}` &nbsp; {date} · {dur}")
            with c2:
                if rid not in st.session_state.pinned_runs:
                    if st.button("Pin tab", key=f"pin_{rid}", use_container_width=True):
                        st.session_state.pinned_runs.insert(0, rid)
                        st.rerun()
                else:
                    st.caption("✓ Pinned")
            with c3:
                st.markdown(f"[GitHub ↗]({run['html_url']})")


# ══════════════════════════════════════════════════════════════════════════════
# TABS 1+ — Individual run tabs
# ══════════════════════════════════════════════════════════════════════════════

for tab_idx, run_id in enumerate(pinned):
    with tabs[tab_idx + 1]:

        run = run_lookup.get(run_id)
        if not run:
            st.warning(f"Could not load run {run_id}")
            continue

        # Refresh live run data
        if run.get("status") != "completed":
            run = get_run(run_id)
            run_lookup[run_id] = run

        # Guard against API error dicts
        if run.get("_error") or "created_at" not in run:
            st.error(f"Could not load run data (ID {run_id}). GitHub API may be temporarily unavailable.")
            if st.button("Remove tab", key=f"rm_{run_id}"):
                st.session_state.pinned_runs.remove(run_id)
                st.rerun()
            continue

        status  = run.get("status", "unknown")
        conc    = run.get("conclusion")
        icon    = status_icon(status, conc)
        name    = run_display_name(run)
        label   = run.get("_workflow_label", "Run")
        dur     = fmt_duration(run["created_at"], run.get("updated_at") if status == "completed" else None)
        run_url = run.get("html_url", "")

        # ── Run header ─────────────────────────────────────────────────────────
        hcol1, hcol2, hcol3 = st.columns([5, 1, 1])
        with hcol1:
            st.subheader(f"{icon} {tab_name_for_run(run_id)}")
            # Show search criteria as inline tags
            inputs = st.session_state.run_inputs_store.get(run_id, {})
            tags = []
            if inputs.get("mode"):        tags.append(f"📋 {inputs['mode']}")
            if inputs.get("region"):      tags.append(f"📍 {inputs['region']}")
            if inputs.get("min_revenue"): tags.append(f"💰 £{int(inputs['min_revenue']):,}+ revenue")
            if inputs.get("notify_email"):tags.append(f"📧 {inputs['notify_email']}")
            if tags:
                st.caption("  ·  ".join(tags))
            else:
                st.caption(f"{label} · {run['created_at'][:10]} · {dur} · {conc or status}")
        with hcol2:
            st.link_button("View on GitHub", run_url, use_container_width=True)
        with hcol3:
            if st.button("🗑 Delete Tab", key=f"close_{run_id}", use_container_width=True,
                         help="Remove this tab. The run itself keeps going on GitHub."):
                st.session_state.pinned_runs.remove(run_id)
                st.rerun()

        st.divider()

        # ── Status card ────────────────────────────────────────────────────────
        if status == "completed" and conc == "success":
            st.success("✅ Run completed successfully")

            artifacts = get_artifacts(run_id)
            excel_arts = [a for a in artifacts if a["name"] in ("pe-sourcing-results", "lift-maintenance-excel")]

            if excel_arts:
                art = excel_arts[0]
                size_kb = art["size_in_bytes"] // 1024
                st.markdown(f"**📊 Excel report ready** — {art['name']} ({size_kb} KB)")
                # Cache the downloaded artifact bytes in session state so
                # clicking "Download" once lets both the save button and
                # the preview render without a second API call.
                cache_key = f"_artifact_bytes_{run_id}"
                if st.button(f"⬇ Download & Preview Excel", key=f"dl_{run_id}", type="primary"):
                    with st.spinner("Downloading from GitHub…"):
                        data = download_artifact(art["id"])
                    if data:
                        st.session_state[cache_key] = data

                if st.session_state.get(cache_key):
                    data = st.session_state[cache_key]
                    st.download_button(
                        label="💾 Save to your computer",
                        data=data,
                        file_name=f"PE_Sourcing_{run['created_at'][:10]}.zip",
                        mime="application/zip",
                        key=f"save_{run_id}",
                    )
                    show_excel_preview(data, run_id)

            # ── Re-enrichment panel ────────────────────────────────────────────
            _show_reenrich_panel(run_id, run, inputs)

            if not excel_arts:
                st.info("No Excel artifact found — it may have expired (30-day retention).")

        elif status == "completed" and conc in ("failure", "timed_out"):
            st.error(f"❌ Run failed ({conc}). [Check logs on GitHub →]({run_url})")
            st.info("GitHub Actions logs will show what went wrong. You can re-trigger from New Search.")

        elif status == "completed" and conc == "cancelled":
            st.warning("🚫 Run was cancelled.")

        elif status in ("queued", "in_progress"):
            # ── Estimated progress bar ─────────────────────────────────────────
            is_deep   = run.get("_workflow_label", "") == "Deep OCR"
            est_mins  = 330 if is_deep else 90          # expected total minutes
            est_secs  = est_mins * 60

            started_str = run.get("run_started_at") or run.get("created_at", "")
            if started_str and status == "in_progress":
                try:
                    started_dt  = datetime.fromisoformat(started_str.replace("Z", "+00:00"))
                    elapsed_sec = (datetime.now(timezone.utc) - started_dt).total_seconds()
                    progress    = min(elapsed_sec / est_secs, 0.97)   # cap at 97% until done
                    elapsed_min = int(elapsed_sec // 60)
                    remain_min  = max(0, int((est_secs - elapsed_sec) // 60))
                    bar_text    = (
                        f"~{elapsed_min} min elapsed · ~{remain_min} min remaining "
                        f"(estimated {est_mins} min total)"
                    )
                except Exception:
                    progress, bar_text = 0.05, "Starting up…"
            elif status == "queued":
                progress, bar_text = 0.02, "Queued — waiting for a GitHub Actions runner…"
            else:
                progress, bar_text = 0.05, "Starting up…"

            st.progress(progress, text=bar_text)
            st.caption(f"Auto-refreshes every 30 s &nbsp;·&nbsp; [Watch live logs on GitHub →]({run_url})")

            # Cancel button
            if st.button("🛑 Cancel this run", key=f"cancel_{run_id}"):
                cancel_run(run_id)
                st.warning("Cancellation requested…")
                time.sleep(3)
                st.cache_data.clear()
                st.rerun()

        # ── Run metadata ───────────────────────────────────────────────────────
        with st.expander("Run details"):
            st.json({
                "run_id":     run_id,
                "status":     status,
                "conclusion": conc,
                "started":    run.get("created_at"),
                "updated":    run.get("updated_at"),
                "duration":   dur,
                "github_url": run_url,
                "workflow":   run.get("_workflow_label", ""),
            })


# ── Auto-refresh if any run is still active ────────────────────────────────────
has_active = any(
    run_lookup.get(rid, {}).get("status") in ("queued", "in_progress")
    for rid in pinned
)
if has_active:
    time.sleep(30)
    st.cache_data.clear()
    st.rerun()
