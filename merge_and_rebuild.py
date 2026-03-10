#!/usr/bin/env python3
"""
merge_and_rebuild.py — Merge OCR accounts + run intelligence layers + rebuild Excel

Steps:
  1. Load lift_maintenance_enriched.json from repo
  2. Patch company_age_years from SQLite sector_cache
  3. Merge OCR accounts data from /tmp/lift_accounts.json
     (replaces estimates with Tier 1 actual figures where available)
  4. Run competitor_map (geographic proximity scoring — no API calls)
  5. Run acquisition_score (5-dimension attractiveness — no API calls)
  6. Patch seller_likelihood into existing sell_intent data (if present)
  7. Rebuild Excel with 6 sheets including all new intelligence
  8. Save to /mnt/outputs/UK_Lift_Maintenance_Companies_March2026.xlsx
"""

import json
import os
import sys
import sqlite3
from pathlib import Path
import datetime

# ── paths — supports local and CI (GitHub Actions) environments ───────────────
# When running on GitHub Actions, the working directory IS the repo root.
# When running locally in the Cowork VM, use the absolute paths below.
# Override any path via environment variables.

_SCRIPT_DIR = Path(__file__).resolve().parent  # wherever this script lives

REPO_DIR      = Path(os.environ.get("REPO_DIR",      str(_SCRIPT_DIR)))
ENRICHED_JSON = Path(os.environ.get("ENRICHED_JSON", str(REPO_DIR / "data/sectors/lift_maintenance_enriched.json")))
OCR_JSON      = Path(os.environ.get("OCR_JSON",      "/tmp/lift_accounts.json"))
DB_PATH       = Path(os.environ.get("DB_PATH",       str(REPO_DIR / "data/companies_house.db")))

# Output: use CI_OUTPUT_PATH env var (set by workflow), else default local path
_default_out  = str(_SCRIPT_DIR.parent / "mnt/outputs/UK_Lift_Maintenance_Companies_March2026.xlsx")
OUT_PATH      = Path(os.environ.get("CI_OUTPUT_PATH", _default_out))
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(REPO_DIR))

# ── 1. Load enriched JSON ──────────────────────────────────────────────────────
print("Loading enriched JSON...")
with open(ENRICHED_JSON) as f:
    data = json.load(f)
print(f"  {len(data)} companies loaded")

# ── 2. Patch company_age_years from DB ────────────────────────────────────────
print("Patching company_age_years from DB...")
try:
    con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    rows = con.execute(
        "SELECT company_number, company_age_years FROM sector_cache WHERE sector='lift_maintenance'"
    ).fetchall()
    con.close()
    age_map = {r[0]: r[1] for r in rows}
    patched = 0
    for r in data:
        cn = r.get("company_number", "")
        if cn in age_map and age_map[cn]:
            r["company_age_years"] = age_map[cn]
            patched += 1
    print(f"  {patched} companies patched with company_age_years")
except Exception as e:
    print(f"  DB patch failed: {e} — falling back to age_years from JSON")

# Fallback: ensure every record has company_age_years set
# (enriched JSON stores this as age_years; DB patches it as company_age_years)
for r in data:
    if not r.get("company_age_years"):
        r["company_age_years"] = r.get("age_years") or 0

# ── 3. Merge OCR accounts data ────────────────────────────────────────────────
print("Merging OCR accounts data...")
if OCR_JSON.exists():
    with open(OCR_JSON) as f:
        ocr_records = json.load(f)
    ocr_map = {r["company_number"]: r for r in ocr_records}
    ocr_merged = 0
    tier1_turnover = 0
    tier1_netassets = 0

    for company in data:
        cn = company.get("company_number", "")
        ocr = ocr_map.get(cn)
        if not ocr or "error" in ocr:
            continue

        # Store raw OCR figures on company
        for field in ("turnover", "operating_profit", "profit_before_tax",
                      "staff_costs", "net_assets", "total_assets",
                      "fixed_assets", "current_assets", "trade_debtors",
                      "total_liabilities", "currency", "accounts_type",
                      "period_end", "filing_date", "figures_extracted"):
            val = ocr.get(field)
            if val is not None:
                company[field] = val

        # FX conversion
        currency = ocr.get("currency", "GBP")
        fx = {"GBP": 1.0, "USD": 0.79, "EUR": 0.86}.get(currency, 1.0)

        turnover = ocr.get("turnover")
        if turnover and turnover > 0:
            gbp_turn = round(turnover * fx)
            company["rev_actual"]    = gbp_turn
            company["rev_low"]       = round(gbp_turn * 0.92)
            company["rev_high"]      = round(gbp_turn * 1.08)
            company["rev_source"]    = "Tier 1 — CH filed accounts (OCR)"
            company["confidence"]    = "Actual (Tier 1)"
            # Also store in revenue_estimate envelope for downstream compatibility
            if "revenue_estimate" not in company or not isinstance(company.get("revenue_estimate"), dict):
                company["revenue_estimate"] = {}
            company["revenue_estimate"]["revenue_mid"] = gbp_turn
            company["revenue_estimate"]["data_tier"]   = "Tier 1 — CH filed accounts (OCR)"
            tier1_turnover += 1
        elif ocr.get("net_assets"):
            na = ocr.get("net_assets", 0) * fx
            company["net_assets_actual"] = round(na)
            company["rev_source"] = "Tier 1B — CH filed accounts (balance sheet only)"
            company["confidence"] = "Actual BS (Tier 1B)"
            tier1_netassets += 1

        ocr_merged += 1

    print(f"  {ocr_merged} OCR records merged")
    print(f"  {tier1_turnover} companies with actual turnover (Tier 1)")
    print(f"  {tier1_netassets} companies with actual net assets (Tier 1B)")
else:
    print(f"  OCR file not found at {OCR_JSON} — using estimates only")

# ── 4. Run competitor_map ─────────────────────────────────────────────────────
print("Running competitor map (no API calls)...")
import competitor_map as cm

index_by_number = {c["company_number"]: c for c in data}
for i, company in enumerate(data):
    if i % 200 == 0:
        print(f"  [{i+1}/{len(data)}] mapping competitors...")
    company["competitor_analysis"] = cm.build_competitor_map(
        target=company,
        all_companies=data,
        index_by_number=index_by_number,
        top_n=10,
    )
print("  Competitor mapping complete")

# ── 5. Run acquisition_score ──────────────────────────────────────────────────
print("Running acquisition attractiveness scoring...")
import acquisition_score as acs

tier_counts = {"Tier 1": 0, "Tier 2": 0, "Tier 3": 0, "Tier 4": 0, "OOS": 0}
for company in data:
    result = acs.acquisition_attractiveness_score(company)
    company["acquisition_attractiveness"] = result
    label = acs.acquisition_tier_label(result["acquisition_score"])
    tier_counts[label] = tier_counts.get(label, 0) + 1

print(f"  Tier 1 (Prime, 80+): {tier_counts['Tier 1']}")
print(f"  Tier 2 (Strong, 65-79): {tier_counts['Tier 2']}")
print(f"  Tier 3 (Bolt-on, 50-64): {tier_counts['Tier 3']}")
print(f"  Tier 4 (Watch, 35-49): {tier_counts['Tier 4']}")
print(f"  OOS (<35): {tier_counts['OOS']}")

# ── 6. Patch seller_likelihood into existing sell_intent data ─────────────────
print("Patching seller likelihood signals...")
from sell_signals import seller_likelihood_score
for company in data:
    e = seller_likelihood_score(company, company.get("company_number", ""))
    si = company.get("sell_intent") or {}
    si["seller_likelihood"]      = e["seller_likelihood"]
    si["seller_likelihood_score"]= e["seller_likelihood_score"]
    si["seller_signals"]         = e["seller_signals"]
    si["seller_signal_flags"]    = e["seller_signal_flags"]
    if "components" not in si:
        si["components"] = {}
    si["components"]["seller_likelihood"] = e
    company["sell_intent"] = si
print("  Seller likelihood patched")

# ── Helper functions ────────────────────────────────────────────────────────────
def fmt_gbp(val):
    if val is None: return "—"
    if val >= 1_000_000: return f"£{val/1_000_000:.1f}m"
    if val >= 1_000: return f"£{val/1_000:.0f}k"
    return f"£{val:.0f}"

def rev_for(r):
    return (r.get("rev_actual") or
            (r.get("revenue_estimate") or {}).get("revenue_mid") or 0)

def rev_label(r):
    src = r.get("rev_source", "")
    if "Tier 1 —" in src or src.startswith("Tier 1"):
        return "Actual"
    if "Tier 1B" in src:
        return "Actual BS"
    return "Estimated"

# ── 7. Build Excel ─────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import datetime

print("Building Excel workbook...")

# Colours
DARK_NAVY   = "0D1B3E"
MID_NAVY    = "1B3A6B"
ACCENT_GOLD = "C9A84C"
LIGHT_GREY  = "F4F6F9"
MID_GREY    = "D9DEE8"
WHITE       = "FFFFFF"
GREEN_PALE  = "E8F5E9"
AMBER_PALE  = "FFF8E1"
RED_PALE    = "FFEBEE"

def hfont(sz=11, bold=True, color=WHITE):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def bfont(sz=9, bold=False, color="222222"):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()
wb.remove(wb.active)

# Sort data by acquisition score desc, then age desc
data.sort(key=lambda x: (
    -(x.get("acquisition_attractiveness") or {}).get("acquisition_score", 0),
    -(x.get("company_age_years") or 0)
))

# ════════════════════════════════════════════════════════════
# SHEET 1: SUMMARY DASHBOARD
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:O1")
t = ws["A1"]
t.value = f"UK LIFT MAINTENANCE — PE DEAL INTELLIGENCE PLATFORM  |  March 2026  |  {len(data)} Companies"
t.font = Font(name="Arial", size=13, bold=True, color=WHITE)
t.fill = fill(DARK_NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

# KPIs
tier1_cos  = sum(1 for r in data if (r.get("acquisition_attractiveness") or {}).get("acquisition_score", 0) >= 80)
tier2_cos  = sum(1 for r in data if 65 <= (r.get("acquisition_attractiveness") or {}).get("acquisition_score", 0) < 80)
with_actual= sum(1 for r in data if rev_label(r) in ("Actual", "Actual BS"))
strong_sell= sum(1 for r in data if (r.get("sell_intent") or {}).get("sell_intent_band") == "Strong")
high_seller= sum(1 for r in data if (r.get("sell_intent") or {}).get("seller_likelihood") == "High")
pe_flagged = sum(1 for r in data if (r.get("competitor_analysis") or {}).get("pe_backed_competitors"))

kpis = [
    (len(data),     "Total Companies"),
    (tier1_cos,     "Tier 1 Targets\n(Acq Score 80+)"),
    (tier2_cos,     "Tier 2 Candidates\n(Score 65-79)"),
    (with_actual,   "With Actual\nAccounts (Tier 1)"),
    (strong_sell,   "Strong Sell\nSignals"),
    (high_seller,   "High Seller\nLikelihood"),
    (pe_flagged,    "PE-Backed\nCompetitors Found"),
]

kpi_cols = ["B","D","F","H","J","L","N"]
kpi_bg   = [DARK_NAVY, "1A5276", "1A5276", "1B4F72", "6E2F2F", "6E2F2F", "4A235A"]
ws.row_dimensions[3].height = 40
ws.row_dimensions[4].height = 24
for idx, ((val, lbl), col_l) in enumerate(zip(kpis, kpi_cols)):
    col_n = ord(col_l) - 64
    c1 = ws.cell(row=3, column=col_n, value=val)
    c1.font = Font(name="Arial", size=20, bold=True, color=WHITE)
    c1.fill = fill(kpi_bg[idx % len(kpi_bg)])
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws.cell(row=4, column=col_n, value=lbl)
    c2.font = Font(name="Arial", size=8, bold=False, color="AAAAAA")
    c2.fill = fill(kpi_bg[idx % len(kpi_bg)])
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[col_l].width = 16

# Top Tier 1 table on summary sheet
ws.row_dimensions[6].height = 20
h = ws["A6"]
h.value = "▶  TOP TIER 1 ACQUISITION TARGETS"
h.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GOLD)
h.fill = fill(DARK_NAVY)

tbl_hdrs = ["Rank", "Company", "Town", "Rev (Actual/Est)", "Acq Score", "Tier",
            "Sell Intent", "Seller Likelihood", "Fragmentation", "PE Competitors"]
for ci, hdr in enumerate(tbl_hdrs, 1):
    c = ws.cell(row=7, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(MID_NAVY)
    c.alignment = Alignment(horizontal="center")

tier1_list = [r for r in data if (r.get("acquisition_attractiveness") or {}).get("acquisition_score", 0) >= 65][:30]
row = 8
for rank, r in enumerate(tier1_list, 1):
    acq   = r.get("acquisition_attractiveness") or {}
    si    = r.get("sell_intent") or {}
    ca    = r.get("competitor_analysis") or {}
    rev   = rev_for(r)
    score = acq.get("acquisition_score", 0)
    bg = GREEN_PALE if score >= 80 else AMBER_PALE if score >= 65 else WHITE
    vals = [
        rank,
        r.get("company_name",""),
        (r.get("registered_office_address") or {}).get("locality") or "",
        f"{fmt_gbp(rev)} ({rev_label(r)})",
        score,
        acq.get("acquisition_tier","")[:20],
        si.get("sell_intent_band",""),
        si.get("seller_likelihood",""),
        ca.get("fragmentation_score",""),
        ", ".join((ca.get("pe_backed_competitors") or [])[:2]) or "—",
    ]
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = bfont(sz=9)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center" if ci != 2 else "left")
    row += 1
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 20
ws.freeze_panes = "A8"

# ════════════════════════════════════════════════════════════
# SHEET 2: ALL COMPANIES (full universe)
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Companies")
ws2.sheet_view.showGridLines = False

hdrs2 = [
    ("Company Name",         30), ("Company No",        12), ("Town",            16),
    ("Postcode",             10), ("SIC",                8), ("Age (Yrs)",         8),
    ("Accts Type",           14), ("Rev (Actual/Est)",  18), ("Rev Source",        14),
    ("Confidence",           12), ("Net Assets",        12), ("Acq Score",         10),
    ("Acq Tier",             22), ("Sell Intent",        10), ("Sell Band",         10),
    ("Seller Likelihood",    16), ("Fragmentation",      13), ("Local Rivals",       12),
    ("PE Competitors",       24), ("Directors",           8), ("Family Co",          9),
    ("Charges (Secured)",    14), ("Employees",           12),
]

for ci, (hdr, _) in enumerate(hdrs2, 1):
    c = ws2.cell(row=1, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for ci, (_, wid) in enumerate(hdrs2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = wid

for row_i, r in enumerate(data, 2):
    acq = r.get("acquisition_attractiveness") or {}
    si  = r.get("sell_intent") or {}
    ca  = r.get("competitor_analysis") or {}
    rev = rev_for(r)
    score = acq.get("acquisition_score", 0)

    charges = r.get("charges") or r.get("outstanding_charges") or {}
    if isinstance(charges, dict):
        ch_outstanding = charges.get("outstanding_charges", 0) or 0
    else:
        ch_outstanding = int(charges) if str(charges).isdigit() else 0

    bs = r.get("bs") or {}
    acct_type = r.get("accounts_type") or bs.get("accounts_type") or ""

    row_data = [
        r.get("company_name",""),
        r.get("company_number",""),
        (r.get("registered_office_address") or {}).get("locality") or "",
        r.get("postcode",""),
        str(r.get("sic1","")) or "",
        r.get("company_age_years") or "",
        acct_type,
        fmt_gbp(rev) if rev else "—",
        rev_label(r),
        r.get("confidence",""),
        fmt_gbp(r.get("net_assets") or r.get("net_assets_actual")),
        score,
        acq.get("acquisition_tier","")[:25] if acq.get("acquisition_tier") else "",
        si.get("sell_intent_score",""),
        si.get("sell_intent_band",""),
        si.get("seller_likelihood",""),
        ca.get("fragmentation_score",""),
        ca.get("competitor_count_local",""),
        ", ".join((ca.get("pe_backed_competitors") or [])[:2]) or "—",
        len(r.get("directors") or []),
        "Yes" if r.get("is_family_company") or r.get("family_business") else "No",
        ch_outstanding,
        r.get("employees_est_band",""),
    ]

    bg = GREEN_PALE if score >= 80 else AMBER_PALE if score >= 65 else WHITE if score >= 50 else LIGHT_GREY
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=row_i, column=ci, value=val)
        c.font = bfont(sz=8)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center")

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdrs2))}1"

# ════════════════════════════════════════════════════════════
# SHEET 3: TIER 1 + TIER 2 TARGETS (detailed)
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Prime Targets")
ws3.sheet_view.showGridLines = False

prime_targets = [r for r in data if (r.get("acquisition_attractiveness") or {}).get("acquisition_score", 0) >= 65]
print(f"  Prime targets (score ≥65): {len(prime_targets)}")

hdrs3 = [
    ("Company Name", 36),     ("Company No", 12),    ("Town", 16),
    ("Age (Yrs)", 8),          ("Accts Type", 14),    ("Revenue", 16),
    ("Rev Source", 14),        ("PBT", 12),           ("Net Assets", 12),
    ("Trade Debtors", 13),     ("Acq Score", 10),     ("Acq Tier", 22),
    ("Sell Intent", 10),       ("Sell Band", 10),     ("Seller Likelihood", 16),
    ("Seller Signals", 10),    ("Fragmentation", 13), ("Local Rivals", 10),
    ("PE Comps", 24),          ("Largest Local Rival", 30),
    ("Directors", 8),          ("Charges", 10),       ("Acq Signals (top 3)", 50),
]

for ci, (hdr, _) in enumerate(hdrs3, 1):
    c = ws3.cell(row=1, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
for ci, (_, wid) in enumerate(hdrs3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = wid

for row_i, r in enumerate(prime_targets, 2):
    acq   = r.get("acquisition_attractiveness") or {}
    si    = r.get("sell_intent") or {}
    ca    = r.get("competitor_analysis") or {}
    rev   = rev_for(r)
    score = acq.get("acquisition_score", 0)
    signals_top3 = " | ".join((acq.get("acquisition_signals") or [])[:3])

    charges = r.get("charges") or r.get("outstanding_charges") or {}
    ch_outstanding = (charges.get("outstanding_charges", 0) if isinstance(charges, dict)
                      else int(charges) if str(charges).isdigit() else 0)
    bs = r.get("bs") or {}
    acct_type = r.get("accounts_type") or bs.get("accounts_type") or ""

    row_data = [
        r.get("company_name",""),
        r.get("company_number",""),
        (r.get("registered_office_address") or {}).get("locality") or "",
        r.get("company_age_years",""),
        acct_type,
        fmt_gbp(rev) if rev else "—",
        rev_label(r),
        fmt_gbp(r.get("profit_before_tax")),
        fmt_gbp(r.get("net_assets") or r.get("net_assets_actual")),
        fmt_gbp(r.get("trade_debtors")),
        score,
        acq.get("acquisition_tier","")[:25] if acq.get("acquisition_tier") else "",
        si.get("sell_intent_score",""),
        si.get("sell_intent_band",""),
        si.get("seller_likelihood",""),
        si.get("seller_likelihood_score",""),
        ca.get("fragmentation_score",""),
        ca.get("competitor_count_local",""),
        ", ".join((ca.get("pe_backed_competitors") or [])[:2]) or "—",
        ca.get("largest_local_competitor","") or "—",
        len(r.get("directors") or []),
        ch_outstanding,
        signals_top3,
    ]

    bg = GREEN_PALE if score >= 80 else AMBER_PALE
    for ci, val in enumerate(row_data, 1):
        c = ws3.cell(row=row_i, column=ci, value=val)
        c.font = bfont(sz=8)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left" if ci in (1,23) else "center",
                                 wrap_text=(ci == 23))
    ws3.row_dimensions[row_i].height = 28

ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(hdrs3))}1"

# ════════════════════════════════════════════════════════════
# SHEET 4: COMPETITOR MAP (top targets)
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Competitor Maps")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:L1")
h1 = ws4["A1"]
h1.value = "COMPETITOR MAP — Top 10 geographic/operational competitors per company (showing top 50 acquisition targets)"
h1.font = hfont(sz=10)
h1.fill = fill(DARK_NAVY)
h1.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 22

hdrs4 = ["Target Company", "Competitor", "Town", "Distance Band",
         "Est Revenue", "Acct Type", "Is PE-Backed", "Is Group",
         "Acq Fit", "Sell Intent Band", "Sell Score"]
for ci, hdr in enumerate(hdrs4, 1):
    c = ws4.cell(row=2, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(MID_NAVY)
    c.alignment = Alignment(horizontal="center")

col_widths4 = [36, 36, 16, 14, 14, 12, 11, 10, 10, 14, 10]
for ci, wid in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(ci)].width = wid

top_for_map = prime_targets[:50]
row = 3
for target in top_for_map:
    ca = target.get("competitor_analysis") or {}
    for comp in (ca.get("competitor_map") or []):
        row_data = [
            target.get("company_name",""),
            comp.get("company_name",""),
            comp.get("town",""),
            comp.get("distance_band",""),
            fmt_gbp(comp.get("estimated_revenue_gbp")),
            comp.get("accounts_type",""),
            "✓" if comp.get("is_pe_backed") else "",
            "✓" if comp.get("is_group_owned") else "",
            comp.get("acquisition_fit",""),
            comp.get("sell_intent_band",""),
            comp.get("sell_intent_score",""),
        ]
        bg = RED_PALE if comp.get("is_pe_backed") else AMBER_PALE if comp.get("distance_band") == "Local" else WHITE
        for ci, val in enumerate(row_data, 1):
            c = ws4.cell(row=row, column=ci, value=val)
            c.font = bfont(sz=8)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
        row += 1

ws4.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════
# SHEET 5: DIRECTORS REGISTER
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Directors Register")
ws5.sheet_view.showGridLines = False

hdrs5 = ["Company", "Company No", "Director Name", "Age", "Tenure (Yrs)",
         "Role", "Nationality", "Sell Signal?", "Acq Score", "Seller Likelihood"]
for ci, hdr in enumerate(hdrs5, 1):
    c = ws5.cell(row=1, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(DARK_NAVY)
    c.alignment = Alignment(horizontal="center")

col_widths5 = [36, 12, 30, 8, 12, 20, 14, 12, 10, 16]
for ci, wid in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = wid

row = 2
for r in data[:500]:  # top 500 by acq score
    acq = r.get("acquisition_attractiveness") or {}
    si  = r.get("sell_intent") or {}
    score = acq.get("acquisition_score", 0)
    for d in (r.get("directors") or []):
        if d.get("resigned"):
            continue
        age_sig = (d.get("age") or 0) >= 60
        ten_sig = (d.get("years_active") or 0) >= 15
        sell_flag = "✓ Age" if age_sig else ("✓ Tenure" if ten_sig else "")
        row_data = [
            r.get("company_name",""),
            r.get("company_number",""),
            d.get("name",""),
            d.get("age",""),
            round(d.get("years_active",0), 1) if d.get("years_active") else "",
            d.get("occupation",""),
            d.get("nationality",""),
            sell_flag,
            score,
            si.get("seller_likelihood",""),
        ]
        bg = AMBER_PALE if (age_sig or ten_sig) else WHITE
        for ci, val in enumerate(row_data, 1):
            c = ws5.cell(row=row, column=ci, value=val)
            c.font = bfont(sz=8)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left" if ci in (1,3,6) else "center")
        row += 1
ws5.freeze_panes = "A2"
ws5.auto_filter.ref = f"A1:{get_column_letter(len(hdrs5))}1"

# ════════════════════════════════════════════════════════════
# SHEET 6: FAMILY BUSINESSES
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Family Businesses")
ws6.sheet_view.showGridLines = False

fam = [r for r in data if r.get("is_family_company") or r.get("family_business")]
print(f"  Family businesses: {len(fam)}")

ws6.merge_cells("A1:M1")
h1 = ws6["A1"]
h1.value = f"FAMILY-OWNED BUSINESSES — {len(fam)} identified (shared director surnames, typically founder-led)"
h1.font = hfont(sz=10)
h1.fill = fill(DARK_NAVY)
h1.alignment = Alignment(horizontal="left", vertical="center")
ws6.row_dimensions[1].height = 22

hdrs6 = ["Company", "Company No", "Town", "Age (Yrs)", "Revenue",
         "Rev Source", "Net Assets", "Acq Score", "Sell Intent", "Sell Band",
         "Seller Likelihood", "Directors", "Charges"]
for ci, hdr in enumerate(hdrs6, 1):
    c = ws6.cell(row=2, column=ci, value=hdr)
    c.font = hfont(sz=9)
    c.fill = fill(MID_NAVY)
    c.alignment = Alignment(horizontal="center")

col_widths6 = [36, 12, 16, 9, 14, 14, 12, 10, 10, 10, 16, 10, 10]
for ci, wid in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(ci)].width = wid

for row_i, r in enumerate(fam, 3):
    acq = r.get("acquisition_attractiveness") or {}
    si  = r.get("sell_intent") or {}
    rev = rev_for(r)
    charges = r.get("charges") or r.get("outstanding_charges") or {}
    ch_out = (charges.get("outstanding_charges", 0) if isinstance(charges, dict) else 0)
    row_data = [
        r.get("company_name",""),
        r.get("company_number",""),
        (r.get("registered_office_address") or {}).get("locality") or "",
        r.get("company_age_years",""),
        fmt_gbp(rev),
        rev_label(r),
        fmt_gbp(r.get("net_assets") or r.get("net_assets_actual")),
        acq.get("acquisition_score",""),
        si.get("sell_intent_score",""),
        si.get("sell_intent_band",""),
        si.get("seller_likelihood",""),
        len(r.get("directors") or []),
        ch_out,
    ]
    bg = GREEN_PALE if acq.get("acquisition_score",0) >= 65 else AMBER_PALE if acq.get("acquisition_score",0) >= 50 else WHITE
    for ci, val in enumerate(row_data, 1):
        c = ws6.cell(row=row_i, column=ci, value=val)
        c.font = bfont(sz=8)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
ws6.freeze_panes = "A3"
ws6.auto_filter.ref = f"A2:{get_column_letter(len(hdrs6))}2"

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(str(OUT_PATH))
print(f"\n✓ Excel saved → {OUT_PATH}")

# ── Summary ────────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"UK Lift Maintenance — Intelligence Platform Rebuild")
print(f"{'='*60}")
print(f"  Companies:          {len(data)}")
print(f"  Tier 1 (Acq 80+):  {tier_counts.get('Tier 1', 0)}")
print(f"  Tier 2 (65-79):     {tier_counts.get('Tier 2', 0)}")
print(f"  With Actual Rev:    {sum(1 for r in data if rev_label(r) == 'Actual')}")
print(f"  With Actual BS:     {sum(1 for r in data if rev_label(r) == 'Actual BS')}")
print(f"  Family-owned:       {len(fam)}")
print(f"  High Seller Lhood:  {sum(1 for r in data if (r.get('sell_intent') or {}).get('seller_likelihood') == 'High')}")
print(f"  Strong Sell Intent: {sum(1 for r in data if (r.get('sell_intent') or {}).get('sell_intent_band') == 'Strong')}")
print(f"{'='*60}")
print(f"\n  Output: {OUT_PATH}")

# ── Email notification ──────────────────────────────────────────────────────────
# Sends completion email with Excel attached. Requires MAIL_USERNAME + MAIL_PASSWORD
# env vars or a .mail_config file. Silently skips if credentials not configured.
try:
    from notify import send_completion_email
    send_completion_email(
        excel_path=OUT_PATH,
        sector="UK Lift Maintenance",
        summary={
            "total":     len(data),
            "tier1":     tier_counts.get("Tier 1", 0),
            "tier2":     tier_counts.get("Tier 2", 0),
            "family":    len(fam),
            "directors": sum(len(r.get("directors", [])) for r in data),
        },
    )
except Exception as e:
    print(f"[notify] Email step failed (non-fatal): {e}")
