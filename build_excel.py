"""
build_excel.py — Generate PE pipeline Excel workbook from enriched JSON

Sheets:
  1.  PE Pipeline          — all companies ranked by acquisition score
  2.  Top 30 Profiles      — detailed per-company intelligence cards
  3.  Director Contacts    — contact intelligence for top companies
  4.  Financials           — revenue/EBITDA estimates and balance sheet
  5.  Bolt-On Analysis     — sector adjacency and roll-up opportunities
  6.  Sell Signals         — sell intent scores and signal breakdown
  7.  Gov. Contracts       — government contract intelligence
  8.  Digital Health       — domain age, LinkedIn, job postings, web presence
  9.  Regulatory Registers — EA, CQC, FCA, ICO, Ofsted, SIA register checks
  10. Summary Stats        — pipeline KPIs
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

import config as cfg

# ── Lazy module imports (only loaded when needed by _normalise) ───────────────
_ch_enrich    = None
_sell_signals = None

def _lazy_ch_enrich():
    global _ch_enrich
    if _ch_enrich is None:
        try:
            import ch_enrich as _m
            _ch_enrich = _m
        except Exception:
            pass
    return _ch_enrich

def _lazy_sell_signals():
    global _sell_signals
    if _sell_signals is None:
        try:
            import sell_signals as _m
            _sell_signals = _m
        except Exception:
            pass
    return _sell_signals


# ── Schema normaliser — adapts enrich_batch and live pipeline dicts ───────────

def _normalise(c: dict) -> dict:
    """
    Ensure every company dict has the fields that build_*() functions expect,
    regardless of whether it came from enrich_batch.py (sector OCR runs) or
    the live run.py pipeline.

    enrich_batch stores:  age_years, incorporation_date, family_business,
                          sic1, age_est, succession_score (top-level), rev_base …
    live pipeline stores: company_age_years, date_of_creation, is_family,
                          sic_codes [], succession {dict}, sell_intent {dict} …
    """
    c = dict(c)  # shallow copy — don't mutate the original

    # ── Basic field aliases ────────────────────────────────────────────────────
    if not c.get("company_age_years"):
        c["company_age_years"] = c.get("age_years", 0)
    if not c.get("date_of_creation"):
        c["date_of_creation"] = c.get("incorporation_date", "")
    if c.get("is_family") is None:
        c["is_family"] = bool(c.get("family_business", False))
    if not c.get("sic_codes"):
        sic1 = c.get("sic1")
        c["sic_codes"] = [sic1] if sic1 else []
    if not c.get("ch_url"):
        cn = c.get("company_number", "")
        if cn:
            c["ch_url"] = (
                f"https://find-and-update.company-information.service.gov.uk/company/{cn}"
            )
    # Ensure director_count is consistent with the directors list
    if not c.get("director_count") and c.get("directors"):
        c["director_count"] = len(c["directors"])

    # ── Normalise director dicts ───────────────────────────────────────────────
    # enrich_batch uses age_est (not age) and has no years_active / occupation
    raw_dirs = c.get("directors", [])
    norm_dirs = []
    for d in raw_dirs:
        nd = dict(d)
        if nd.get("age") is None:
            nd["age"] = nd.get("age_est")
        if nd.get("years_active") is None:
            appt_year = (nd.get("appointed") or "")[:4]
            try:
                nd["years_active"] = round(2026 - int(appt_year), 1) if len(appt_year) == 4 else 0
            except (ValueError, TypeError):
                nd["years_active"] = 0
        if not nd.get("occupation"):
            nd["occupation"] = nd.get("role", "")
        norm_dirs.append(nd)
    c["directors"] = norm_dirs

    # ── Succession dict ────────────────────────────────────────────────────────
    if not c.get("succession"):
        che = _lazy_ch_enrich()
        if che:
            try:
                ss = che.succession_score(norm_dirs)
                c["succession"] = ss
            except Exception:
                pass
        if not c.get("succession"):
            # Minimal fallback from batch top-level fields
            oldest = c.get("oldest_director_age") or max(
                (d["age"] for d in norm_dirs if d.get("age")), default=0
            )
            avg = oldest  # rough proxy
            c["succession"] = {
                "total":   c.get("succession_score", 0),
                "max_age": oldest or None,
                "avg_age": avg or None,
            }

    # ── Dealability ────────────────────────────────────────────────────────────
    if not c.get("dealability"):
        charges = c.get("charges", {})
        outstanding = charges.get("outstanding_charges", 0)
        sc = 0
        # Clean charge register signal
        if outstanding == 0:
            sc += 3
        elif outstanding <= 2:
            sc += 1
        # Governance hire proxy: any director with a known occupation/role
        if any(d.get("occupation") or d.get("role") for d in norm_dirs):
            sc += 1
        c["dealability"] = {
            "score":        min(sc, 20),
            "signals":      [],
            "signal_count": 0,
            "data_tier":    "Tier 4 — derived from charges",
        }

    # ── Acquisition score ──────────────────────────────────────────────────────
    if c.get("acquisition_score") is None:
        che = _lazy_ch_enrich()
        if che:
            try:
                pe  = bool(c.get("pe_backed"))
                acq = che.acquisition_score(
                    int(c.get("company_age_years", 0)),
                    c["succession"],
                    pe,
                    c["dealability"],
                    c.get("charges", {}),
                )
                c["acquisition_score"] = acq["total"]
                c["acquisition_grade"] = che.grade(acq["total"])
                c["acq_components"]    = acq
            except Exception:
                pass
        if c.get("acquisition_score") is None:
            c["acquisition_score"] = 0
            c["acquisition_grade"] = ""
            c["acq_components"]    = {}

    # ── Sell Intent (offline — no API calls) ───────────────────────────────────
    if not c.get("sell_intent"):
        _ss = _lazy_sell_signals()
        if _ss:
            try:
                a = _ss.age_tenure_score(norm_dirs)
                b = _ss.structure_score(norm_dirs)
                d = _ss.maturity_score(int(c.get("company_age_years", 0)))
                # Operational stress proxy: use outstanding charges (no API)
                outstanding = c.get("charges", {}).get("outstanding_charges", 0)
                stress_sc = min(15, outstanding * 5)
                total = min(a["score"] + b["score"] + stress_sc + d["score"], 100)
                if   total >= 70: band = "Strong"
                elif total >= 50: band = "Moderate"
                elif total >= 30: band = "Weak"
                else:             band = "Low"
                e = _ss.seller_likelihood_score(c, c.get("company_number", ""))
                c["sell_intent"] = {
                    "sell_intent_score":       total,
                    "sell_intent_band":        band,
                    "sell_signals":            a["signals"] + b["signals"] + d["signals"],
                    "signal_count":            len(a["signals"]) + len(b["signals"]) + len(d["signals"]),
                    "seller_likelihood":       e.get("seller_likelihood", "Low"),
                    "seller_likelihood_score": e.get("seller_likelihood_score", 0),
                    "seller_signals":          e.get("seller_signals", []),
                    "components": {
                        "age_tenure":         a,
                        "business_structure": b,
                        "company_maturity":   d,
                        "operational_stress": {"score": stress_sc, "signals": [], "data_tier": "Tier 4 — charges proxy"},
                        "seller_likelihood":  e,
                    },
                    "data_tier": "Tier 4 — offline derived (no operational-stress API call)",
                }
            except Exception:
                c["sell_intent"] = {}
        else:
            c["sell_intent"] = {}

    # ── Estimated employees ────────────────────────────────────────────────────
    if c.get("estimated_employees") is None:
        bs  = c.get("bs") or c.get("financials", {}).get("balance_sheet", {}) or {}
        fin = c.get("financials") or {}
        rev = (fin.get("revenue_estimate") or {}).get("revenue_base") or c.get("rev_base")
        emp = bs.get("total_employees")
        if emp and emp > 0:
            c["estimated_employees"]        = int(emp)
            c["estimated_employees_source"] = "Tier 1 — filed accounts"
        else:
            sc = bs.get("staff_costs")
            if sc and sc > 0:
                c["estimated_employees"]        = max(1, round(sc / 35_000))
                c["estimated_employees_source"] = "Tier 4 — staff costs ÷ £35k"
            elif rev and rev > 0:
                c["estimated_employees"]        = max(1, round(rev / 80_000))
                c["estimated_employees_source"] = "Tier 4 — revenue ÷ £80k"

    # ── EBITDA low / high (derive from base using sector margins) ──────────────
    # Only needed when we have an ebitda_base but no range (typical for batch data)
    if c.get("ebitda_base") and not c.get("ebitda_low"):
        try:
            import config as _cfg
            margin_low  = getattr(_cfg, "EBITDA_MARGIN_LOW",  0.10)
            margin_base = getattr(_cfg, "EBITDA_MARGIN_BASE", 0.15)
            margin_high = getattr(_cfg, "EBITDA_MARGIN_HIGH", 0.20)
            eb = c["ebitda_base"]
            # Scale low/high relative to how base relates to its margin
            # (preserves ratio even when base came from a different model)
            if margin_base > 0:
                rev_implied = eb / margin_base
                c["ebitda_low"]  = round(rev_implied * margin_low)
                c["ebitda_high"] = round(rev_implied * margin_high)
            else:
                c["ebitda_low"]  = round(eb * 0.70)
                c["ebitda_high"] = round(eb * 1.35)
        except Exception:
            c["ebitda_low"]  = round(c["ebitda_base"] * 0.70)
            c["ebitda_high"] = round(c["ebitda_base"] * 1.35)

    # ── Normalise financials dict so build_financials() sheet can read it ──────
    # build_financials() reads c["financials"]["revenue_estimate"] etc.
    # For batch data these live at the top level — promote them.
    if not c.get("financials"):
        bs_raw = c.get("bs") or {}
        c["financials"] = {
            "revenue_estimate": {
                "revenue_low":    c.get("rev_low"),
                "revenue_base":   c.get("rev_base"),
                "revenue_high":   c.get("rev_high"),
                "confidence":     c.get("confidence", ""),
                "formula":        "Tier 4 — " + ", ".join(c.get("models_used", ["Location Model"])),
            },
            "ebitda_estimate": {
                "ebitda_low":  c.get("ebitda_low"),
                "ebitda_base": c.get("ebitda_base"),
                "ebitda_high": c.get("ebitda_high"),
            },
            "balance_sheet": {
                "accounts_type":  bs_raw.get("accounts_type", ""),
                "period_end":     bs_raw.get("period_end", ""),
                "net_assets":     bs_raw.get("net_assets"),
                "total_assets":   bs_raw.get("total_assets"),
                "total_employees":bs_raw.get("total_employees"),
                "staff_costs":    bs_raw.get("staff_costs"),
            },
            "balance_sheet_ratios": {
                "net_assets": bs_raw.get("net_assets"),
            },
            "charges": c.get("charges", {}),
        }

    return c

# ── Colour palette ─────────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E75B6"
WHITE  = "FFFFFF"
ALT    = "EBF3FB"
GREEN  = "E2EFDA"
RED    = "FCE4D6"
AMBER  = "FFEB9C"
GREY   = "F2F2F2"
DKGREY = "404040"
PURPLE = "7030A0"
TEAL   = "00B0F0"

THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def score_fill(score):
    if score >= 80: return fill("375623")   # dark green = Prime
    if score >= 65: return fill("70AD47")   # green = High
    if score >= 50: return fill(AMBER)      # amber = Medium
    return fill("FF7070")                    # red = Intel only

def score_font_color(score):
    return WHITE if score >= 65 else "000000"

def sell_intent_fill(band: str):
    mapping = {
        "Strong":   "375623",
        "Moderate": "70AD47",
        "Weak":     AMBER,
        "Low":      "D9D9D9",
    }
    return fill(mapping.get(band, "D9D9D9"))

def sell_intent_font(band: str):
    return WHITE if band in ("Strong", "Moderate") else "000000"

def digital_band_fill(band: str):
    mapping = {
        "Mature":        "375623",
        "Adequate":      "70AD47",
        "Below Average": AMBER,
        "Poor":          "FF7070",
    }
    return fill(mapping.get(band, "D9D9D9"))

def cell(ws, row, col, value, bg=None, fg="000000", bold=False,
         align="left", wrap=False, size=9, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        # bg may already be a PatternFill object (e.g. from sell_intent_fill)
        # or a plain hex string — handle both
        c.fill = bg if isinstance(bg, PatternFill) else fill(bg)
    c.font = Font(name="Arial", size=size, bold=bold, color=fg or "000000")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = THIN
    return c

def title_row(ws, row, cols_span, text, bg=NAVY, size=13):
    ws.merge_cells(f"A{row}:{get_column_letter(cols_span)}{row}")
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = Font(name="Arial", bold=True, size=size, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")

def sub_row(ws, row, cols_span, text, bg=BLUE):
    ws.merge_cells(f"A{row}:{get_column_letter(cols_span)}{row}")
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = Font(name="Arial", italic=True, size=9, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")


# ── Sheet 1: Full pipeline ────────────────────────────────────────────────────

PIPELINE_COLS = [
    ("Rank", 5), ("Reg. No.", 12), ("Company Name", 48), ("Sector ✓", 13), ("Incorp.", 11),
    ("Age", 7), ("Dirs", 5), ("Max Age", 8), ("Avg Age", 8),
    ("Succ.", 7), ("Deal.", 7), ("Acq. Score", 10), ("Grade", 10),
    ("Sell Intent", 10), ("SI Band", 10),
    ("PE", 5), ("Family", 6),
    ("Scale", 7), ("Market", 7), ("Own./Succ.", 10), ("Dealability", 11),
    ("Charges", 8), ("Contracts", 9), ("Digital", 9), ("Accreds", 8), ("SIC", 22),
    ("Director", 28), ("Email", 36), ("Email Conf.", 12), ("Director LinkedIn", 40),
    ("Nearest Competitor", 35), ("Competitor 2", 32), ("Competitor 3", 32),
    # ── Financial intelligence ────────────────────────────────────────────────
    ("Employees", 11), ("Emp. Source", 22),
    ("Rev. Low £", 13), ("Rev. Base £", 13), ("Rev. High £", 13),
    ("Rev. Trend", 12), ("EBITDA £", 13), ("Rev. Conf.", 10),
    ("Yr1 Period", 11), ("Yr1 Accts", 20),
    ("Yr2 Period", 11), ("Yr2 Accts", 20),
    ("Yr3 Period", 11), ("Yr3 Accts", 20),
]

# Column header tooltips — explain the formula or data source for each column
PIPELINE_COL_NOTES = [
    "Rank: ordered 1→N by Acquisition Score (descending).",
    "Reg. No.: Companies House registration number. Source: CH API.",
    "Company Name: Registered name at Companies House.",
    "Sector ✓: Web-verified sector match.\n"
    "  Confirmed (green)   — company name or website clearly mentions the sector keywords.\n"
    "  Likely (yellow)     — partial keyword match on website or name stem.\n"
    "  Uncertain (red)     — website found but no sector keywords detected.\n"
    "  Unverified (grey)   — no website; sector match based on SIC code only.\n"
    "  Source: digital_health.py → sector_relevance_score()",
    "Incorp.: Year of incorporation from Companies House.",
    "Age: Current year − incorporation year. Used in Scale & Financial score.",
    "Dirs: Number of active directors at Companies House.",
    "Max Age: Age of the oldest active director (from CH API date-of-birth data).",
    "Avg Age: Mean age of all active directors with known date-of-birth.",
    "Succ. (Succession Score 0–100): Composite of director age + tenure risk.\n"
    "  High age (≥65) → high score; single director → high concentration risk.",
    "Deal. (Dealability Score 0–20): Signals from CH API:\n"
    "  +charges filed, +director resignations, +dormant subsidiaries,\n"
    "  +group structure simplicity. Raw 0–20 then converted to 0–100 weight.",
    "Acq. Score (0–100): Weighted acquisition attractiveness.\n"
    "  = Scale×0.30 + Market×0.20 + Ownership/Succession×0.30 + Dealability×0.20\n"
    "  Source: ch_enrich.py → acquisition_score()",
    "Grade: Banding of Acquisition Score.\n"
    "  Prime = 80–100 | High = 65–79 | Medium = 50–64 | Intelligence Only = <50",
    "Sell Intent (0–100): Composite exit-readiness signal.\n"
    "  A. Age & Tenure (0–40 pts): founder age + long tenure\n"
    "  B. Business Structure (0–25 pts): single director, no subsidiaries\n"
    "  C. Operational Stress (0–20 pts): charges, cashflow stress\n"
    "  D. Market Maturity (0–15 pts): sector cyclicality, age of firm\n"
    "  Source: sell_signals.py → sell_intent_score()",
    "SI Band: Sell Intent banding.\n"
    "  Hot = 70–100 | Warm = 45–69 | Watchlist = 25–44 | Cold = <25",
    "PE: ✓ if company shows signs of PE backing (investor in name, group structure).\n"
    "  PE-backed companies are deprioritised as acquisition targets.",
    "Family: ✓ if likely family-owned (shared surname directors, long tenures, small board).\n"
    "  Family businesses score higher on succession dimension.",
    "Scale (weighted 0–30): Scale & Financial dimension of Acquisition Score.\n"
    "  Based on company age as maturity proxy (age ≥25 → 90/100 raw × 0.30).\n"
    "  Contribution: Scale_raw × 0.30",
    "Market (weighted 0–20): Market Attractiveness dimension.\n"
    "  Fixed sector-level fragmentation score from SIC discovery × 0.20.\n"
    "  Contribution: MARKET_ATTRACTIVENESS_SCORE × 0.20",
    "Own./Succ. (weighted 0–30): Ownership & Succession dimension.\n"
    "  = (PE_independence_pts(0/40) + succession_score(0–100)) / 1.4 × 0.30\n"
    "  PE-independent companies start with 40 pts; succession risk adds more.",
    "Dealability (weighted 0–20): Dealability Signals dimension.\n"
    "  = (dealability_raw(0–20) / 20) × 100 × 0.20\n"
    "  Source: ch_enrich.py → dealability_score()",
    "Charges: Number of outstanding charges (mortgages/debentures) at Companies House.\n"
    "  High charges may indicate leveraged balance sheet.",
    "Contracts: Government contracts found on Contracts Finder (contracts.service.gov.uk).\n"
    "  Count of contracts awarded in last 3 years.",
    "Digital (0–100): Digital health score.\n"
    "  Composite of: domain age, website presence, LinkedIn activity, job postings.\n"
    "  Source: digital_health.py",
    "Accreds: Number of industry accreditations found (e.g. CHAS, SafeContractor, ISO).\n"
    "  Source: accreditations.py — checks regulatory and trade body registers.",
    "SIC: SIC codes registered at Companies House for this company.",
    "Director: Name of the first active director (for email outreach context).\n"
    "  Source: Companies House director records.",
    "Email: Best inferred director email address for outbound outreach.\n"
    "  Green  = DNS-verified (MX record confirmed, domain accepts mail).\n"
    "  Yellow = Inferred pattern (domain confirmed but email not directly found).\n"
    "  Grey   = No email found.\n"
    "  Source: ch_contacts.py — pattern inference + Disify DNS verification.",
    "Email Conf.: Email confidence level.\n"
    "  High   = Found directly on website or company directory.\n"
    "  Medium = Common pattern (firstname.lastname@domain) with DNS-valid domain.\n"
    "  Low    = Inferred pattern, domain unconfirmed.\n"
    "  Source: ch_contacts.py",
    "Director LinkedIn: LinkedIn profile URL of the director or company page.\n"
    "  Source: ch_contacts.py + digital_health.py (website scrape for LinkedIn links).",
    "Nearest Competitor: Closest competitor company name with distance in miles.\n"
    "  Format: COMPANY NAME (X.X mi)  — sorted by haversine distance from registered postcode.\n"
    "  Source: competitor_map.py + pgeocode lat/lon lookup.",
    "Competitor 2: Second-nearest competitor company name with distance in miles.",
    "Competitor 3: Third-nearest competitor company name with distance in miles.",
    # ── Financial intelligence notes ──────────────────────────────────────────
    "Employees: Best available employee count.\n"
    "  Tier 1 — from filed Companies House accounts (when disclosed).\n"
    "  Tier 4 — estimated from staff costs ÷ £35k avg salary, or revenue ÷ £80k per head.\n"
    "  UK SMEs filing Total Exemption accounts do not disclose headcount publicly.",
    "Emp. Source: Data reliability tier for the employee figure.\n"
    "  Tier 1 = Companies House filing  |  Tier 4 = Derived estimate.",
    "Rev. Low £: Low end of revenue estimate range.\n"
    "  Source: ch_financials.py PE triangulation — Staff Cost, Net Asset, Location, Employee models.",
    "Rev. Base £: Central revenue estimate (most likely scenario).\n"
    "  Source: ch_financials.py PE triangulation — weighted blend of available models.",
    "Rev. High £: High end of revenue estimate range.",
    "Rev. Trend: Year-on-year revenue direction based on accounts history.\n"
    "  ↑ = growing  |  ↓ = declining  |  → = flat  |  ? = insufficient data.\n"
    "  Derived from net assets trend across the last 3 annual filings.",
    "EBITDA £: Estimated EBITDA at base revenue.\n"
    "  Source: Sector-average EBITDA margin applied to Rev. Base.",
    "Rev. Conf.: Revenue estimate confidence level.\n"
    "  HIGH = 3+ financial models available (actual accounts data).\n"
    "  MEDIUM = 2 models.  LOW = 1 model (estimates only).",
    "Yr1 Period: Period end date of the most recent annual accounts filing.",
    "Yr1 Accts: Accounts type for the most recent filing (e.g. Total Exemption, Full).",
    "Yr2 Period: Period end date of the second most recent annual accounts filing.",
    "Yr2 Accts: Accounts type for the second most recent filing.",
    "Yr3 Period: Period end date of the third most recent annual accounts filing.",
    "Yr3 Accts: Accounts type for the third most recent filing.",
]

def build_pipeline(wb, companies):
    ws = wb.active
    ws.title = "PE Pipeline"
    n = len(PIPELINE_COLS)

    title_row(ws, 1, n, f"{cfg.SECTOR_LABEL}  —  PE Pipeline  ({len(companies)} companies)  |  March 2026")
    sub_row(ws, 2, n,
            f"Acquisition Score = Scale(30%) + Market(20%) + Ownership/Succession(30%) + Dealability(20%)  |  "
            f"Sell Intent = Age/Tenure(40) + Structure(25) + Stress(20) + Maturity(15)  |  Source: Companies House API")

    ws.row_dimensions[3].height = 36
    for ci, (label, width) in enumerate(PIPELINE_COLS, 1):
        cell(ws, 3, ci, label, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
        # Add formula/source tooltip as a cell comment
        if ci <= len(PIPELINE_COL_NOTES):
            note_text = PIPELINE_COL_NOTES[ci - 1]
            cmt = Comment(note_text, "V² Pipeline")
            cmt.width  = 320
            cmt.height = max(60, note_text.count("\n") * 18 + 40)
            ws.cell(row=3, column=ci).comment = cmt

    for i, c in enumerate(companies, 1):
        row  = i + 3
        bg   = ALT if i % 2 == 0 else None
        acq  = c["acquisition_score"]
        comp = c.get("acq_components", {})
        ss   = c.get("succession", {})
        deal = c.get("dealability", {})
        ch   = c.get("charges", {})
        si   = c.get("sell_intent", {})
        gc   = c.get("government_contracts", {})
        dh   = c.get("digital_health", {})
        ac   = c.get("accreditations", {})

        cell(ws, row, 1,  i,                              bg=bg, align="center", bold=True)
        cell(ws, row, 2,  c["company_number"],            bg=bg)
        cell(ws, row, 3,  c["company_name"],              bg=bg)

        # ── Col 4: Sector web-verification ────────────────────────────────────
        srl = dh.get("sector_relevance_label", "Unverified")
        srs = dh.get("sector_relevance_score", 0)
        sector_fill_map = {
            "Confirmed":  fill(GREEN),
            "Likely":     fill("FFF2CC"),
            "Uncertain":  fill("FFD6D6"),
            "Unverified": fill("EEEEEE"),
        }
        cx4 = ws.cell(row=row, column=4, value=srl)
        cx4.fill      = sector_fill_map.get(srl, fill("EEEEEE"))
        cx4.font      = Font(name="Arial", size=9, bold=True)
        cx4.alignment = Alignment(horizontal="center", vertical="center")
        cx4.border    = THIN
        if srs:
            cx4.comment = Comment(f"Sector match score: {srs}/100\n"
                                  + "\n".join(dh.get("sector_match_signals", [])),
                                  "V² Pipeline")

        cell(ws, row, 5,  (c.get("date_of_creation") or "")[:4], bg=bg, align="center")
        cell(ws, row, 6,  c.get("company_age_years", 0), bg=bg, align="center")
        cell(ws, row, 7,  c.get("director_count", 0),    bg=bg, align="center")
        cell(ws, row, 8,  ss.get("max_age") or "-",      bg=bg, align="center")
        cell(ws, row, 9,  ss.get("avg_age") or "-",      bg=bg, align="center")
        cell(ws, row, 10, ss.get("total", 0),            bg=bg, align="center")
        cell(ws, row, 11, deal.get("score", 0),          bg=bg, align="center")

        # Acquisition score (cols 12–13)
        for col in (12, 13):
            val = acq if col == 12 else c.get("acquisition_grade", "")
            cx  = ws.cell(row=row, column=col, value=val)
            cx.fill      = score_fill(acq)
            cx.font      = Font(name="Arial", size=9, bold=True, color=score_font_color(acq))
            cx.alignment = Alignment(horizontal="center", vertical="center")
            cx.border    = THIN

        # Sell Intent Score (cols 14–15)
        si_score = si.get("sell_intent_score")
        si_band  = si.get("sell_intent_band", "")
        for col in (14, 15):
            val = si_score if col == 14 else si_band
            cx  = ws.cell(row=row, column=col, value=val if val is not None else "-")
            cx.fill      = sell_intent_fill(si_band) if si_band else fill(GREY)
            cx.font      = Font(name="Arial", size=9, bold=True, color=sell_intent_font(si_band))
            cx.alignment = Alignment(horizontal="center", vertical="center")
            cx.border    = THIN

        cell(ws, row, 16, "⚠" if c.get("pe_backed") else "-",
             bg=RED if c.get("pe_backed") else bg, align="center")
        cell(ws, row, 17, "✓" if c.get("is_family") else "-",
             bg=GREEN if c.get("is_family") else bg, align="center")

        cell(ws, row, 18, comp.get("scale_financial", 0),       bg=bg, align="center")
        cell(ws, row, 19, comp.get("market_attractiveness", 0), bg=bg, align="center")
        cell(ws, row, 20, comp.get("ownership_succession", 0),  bg=bg, align="center")
        cell(ws, row, 21, comp.get("dealability", 0),           bg=bg, align="center")
        cell(ws, row, 22, ch.get("outstanding_charges", "-"),   bg=bg, align="center")
        # Contracts found
        cf = gc.get("contracts_found")
        cell(ws, row, 23, cf if cf is not None else "-",        bg=GREEN if cf else bg, align="center")
        # Digital score
        ds = dh.get("digital_health_score")
        cell(ws, row, 24, ds if ds is not None else "-",        bg=bg, align="center")
        # Accreditation count
        ac_n = ac.get("accreditation_count")
        cell(ws, row, 25, ac_n if ac_n is not None else "-",    bg=GREEN if ac_n else bg, align="center")
        cell(ws, row, 26, ", ".join(c.get("sic_codes", [])),    bg=bg, size=8)

        # ── Director email + LinkedIn (cols 27–29) ────────────────────────────
        contacts  = c.get("contacts", {})
        directors = c.get("directors", [])

        # Best verified email from contacts enrichment
        best_email      = contacts.get("best_email", "")
        email_conf      = contacts.get("email_confidence", "")
        email_verified  = contacts.get("email_dns_valid", False)

        # Director LinkedIn (from contacts or digital_health social signals)
        dir_linkedin = (
            contacts.get("linkedin_url")
            or dh.get("linkedin_url")        # company LinkedIn from website
            or ""
        )

        # Director name (first active director for context)
        dir_name = ""
        for d in directors:
            if d.get("resigned_on") is None and d.get("name"):
                dir_name = d["name"].title()
                break

        # Email cell — green if DNS-verified, amber if inferred, grey if missing
        email_bg = GREEN if email_verified else (fill("FFF2CC") if best_email else fill("EEEEEE"))
        cell(ws, row, 27, dir_name,    bg=bg, size=8)
        cell(ws, row, 28, best_email,  bg=email_bg, size=8)
        cell(ws, row, 29, email_conf,  bg=email_bg, align="center", size=8)
        cell(ws, row, 30, dir_linkedin, bg=bg, size=8)

        # ── Cols 31–33: Nearest 3 competitors ─────────────────────────────────
        comp_map = (c.get("competitor_analysis") or {}).get("competitor_map", [])
        for ci_off, comp_entry in enumerate(comp_map[:3]):
            cname = comp_entry.get("company_name", "")
            miles = comp_entry.get("distance_miles")
            if miles is not None:
                label = f"{cname} ({miles:.1f} mi)"
            else:
                band = comp_entry.get("distance_band", "")
                label = f"{cname} ({band})" if cname else ""
            is_pe = comp_entry.get("is_pe_backed", False)
            comp_bg = fill("FFD6D6") if is_pe else (ALT if i % 2 == 0 else None)
            cell(ws, row, 31 + ci_off, label, bg=comp_bg, size=8, wrap=False)

        # ── Cols 34–46: Financial intelligence ────────────────────────────────
        emp       = c.get("estimated_employees")
        emp_src   = c.get("estimated_employees_source", "")
        # Revenue: try nested financials dict first (live pipeline),
        # then fall back to flat top-level keys (enrich_batch / legacy format)
        fin       = c.get("financials") or {}
        rev_est   = fin.get("revenue_estimate") or {}
        ebitda_est= fin.get("ebitda_estimate") or {}
        rev_low   = rev_est.get("revenue_low")  or c.get("rev_low")
        rev_base  = rev_est.get("revenue_base") or c.get("rev_base")
        rev_high  = rev_est.get("revenue_high") or c.get("rev_high")
        ebitda    = ebitda_est.get("ebitda_base") or c.get("ebitda_base")
        conf      = rev_est.get("confidence") or c.get("confidence", "")
        hist      = c.get("accounts_history") or fin.get("accounts_history") or []

        # Col 34 — Employees (green if Tier 1, amber if estimated)
        emp_bg = GREEN if emp_src.startswith("Tier 1") else (fill("FFF2CC") if emp else bg)
        cell(ws, row, 34, emp if emp is not None else "-", bg=emp_bg, align="center", size=9,
             bold=emp_src.startswith("Tier 1"))

        # Col 35 — Employee source
        cell(ws, row, 35, emp_src, bg=bg, size=7)

        # Cols 36–38 — Revenue Low / Base / High
        def rev_str(v):
            return f"£{v:,.0f}" if v else "-"
        cell(ws, row, 36, rev_str(rev_low),  bg=bg, align="right", size=9)
        cx_rev = ws.cell(row=row, column=37, value=rev_str(rev_base))
        cx_rev.font      = Font(name="Arial", size=9, bold=True)
        cx_rev.alignment = Alignment(horizontal="right", vertical="center")
        cx_rev.border    = THIN
        cx_rev.fill      = fill("E8F4FD") if rev_base else (bg or fill("FFFFFF"))
        cell(ws, row, 38, rev_str(rev_high), bg=bg, align="right", size=9)

        # Col 39 — Revenue trend (derived from accounts history net_assets if available,
        #           otherwise from est revenue confidence direction indicator)
        trend_symbol = "?"
        if len(hist) >= 2:
            na_vals = [h.get("net_assets") for h in hist if h.get("net_assets") is not None]
            if len(na_vals) >= 2:
                diff = na_vals[0] - na_vals[-1]   # most recent vs oldest
                pct  = diff / abs(na_vals[-1]) * 100 if na_vals[-1] else 0
                if pct > 5:
                    trend_symbol = "↑"
                elif pct < -5:
                    trend_symbol = "↓"
                else:
                    trend_symbol = "→"
        trend_bg = fill("E2EFDA") if trend_symbol == "↑" else (
                   fill("FFD6D6") if trend_symbol == "↓" else
                   fill("FFF2CC") if trend_symbol == "→" else
                   (fill(bg) if isinstance(bg, str) else bg) if bg else fill("EEEEEE"))
        trend_color = "1A5C2C" if trend_symbol == "↑" else (
                      "7B0000" if trend_symbol == "↓" else "7B5B00")
        cx_tr = ws.cell(row=row, column=39, value=trend_symbol)
        cx_tr.fill      = trend_bg if trend_bg else fill("EEEEEE")
        cx_tr.font      = Font(name="Arial", size=12, bold=True, color=trend_color)
        cx_tr.alignment = Alignment(horizontal="center", vertical="center")
        cx_tr.border    = THIN

        # Col 40 — EBITDA estimate
        cell(ws, row, 40, rev_str(ebitda), bg=bg, align="right", size=9)

        # Col 41 — Revenue confidence
        conf_bg = fill("E2EFDA") if conf == "HIGH" else (
                  fill("FFF2CC") if conf == "MEDIUM" else
                  fill("FFD6D6") if conf == "LOW" else bg)
        cell(ws, row, 41, conf, bg=conf_bg, align="center", size=8, bold=(conf == "HIGH"))

        # Cols 42–47 — Last 3 years of accounts filings (period + type)
        for yr_idx in range(3):
            base_col = 42 + yr_idx * 2
            if yr_idx < len(hist):
                h = hist[yr_idx]
                period = (h.get("period_end") or "")[:7]   # YYYY-MM
                acc_type = h.get("accounts_type", "")
                cell(ws, row, base_col,     period,   bg=bg, align="center", size=8)
                cell(ws, row, base_col + 1, acc_type, bg=bg, size=7)
            else:
                cell(ws, row, base_col,     "—", bg=bg, align="center", size=8, fg="AAAAAA")
                cell(ws, row, base_col + 1, "—", bg=bg, size=7, fg="AAAAAA")

    ws.freeze_panes = "E4"   # freeze cols A-D (Rank, Reg, Name, Sector ✓)
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(companies)+3}"


# ── Sheet 2: Top 30 profiles ──────────────────────────────────────────────────

def build_top30(wb, companies):
    ws  = wb.create_sheet("Top 30 Profiles")
    top = companies[:30]
    n   = 14

    title_row(ws, 1, n, "TOP 30 PE ACQUISITION TARGETS — INTELLIGENCE PROFILES")

    row = 2
    for rank, c in enumerate(top, 1):
        acq  = c["acquisition_score"]
        ss   = c.get("succession", {})
        deal = c.get("dealability", {})
        comp = c.get("acq_components", {})
        ch   = c.get("charges", {})

        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
        hdr = ws.cell(row=row, column=1,
                      value=f"#{rank}  {c['company_name']}  |  Reg: {c['company_number']}"
                            f"  |  Score: {acq}  |  {c.get('acquisition_grade','')}")
        hdr.fill      = score_fill(acq)
        hdr.font      = Font(name="Arial", bold=True, size=10, color=score_font_color(acq))
        hdr.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Metrics grid
        meta_labels = ["Incorp.", "Age", "Directors", "Max Dir. Age",
                       "Succ. Score", "Deal. Score", "PE Backed", "Charges",
                       "Scale", "Market", "Own./Succ.", "Dealability", "Family", "SIC"]
        meta_vals = [
            (c.get("date_of_creation") or "")[:4],
            f"{c.get('company_age_years',0)} yrs",
            c.get("director_count", 0),
            ss.get("max_age") or "N/A",
            ss.get("total", 0),
            deal.get("score", 0),
            "YES ⚠" if c.get("pe_backed") else "No",
            ch.get("outstanding_charges", "N/A"),
            comp.get("scale_financial", 0),
            comp.get("market_attractiveness", 0),
            comp.get("ownership_succession", 0),
            comp.get("dealability", 0),
            "YES" if c.get("is_family") else "No",
            ", ".join(c.get("sic_codes", [])),
        ]
        for ci, (lbl, val) in enumerate(zip(meta_labels, meta_vals), 1):
            cell(ws, row,   ci, lbl, bg=BLUE, fg=WHITE, bold=True, align="center", size=8)
            cell(ws, row+1, ci, val, bg=ALT,  align="center", size=9)
        row += 2

        # Succession detail
        if ss.get("total"):
            ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
            sh = ws.cell(row=row, column=1,
                         value=f"  Succession Detail  |  Age Score: {ss.get('age_score',0)}  "
                               f"|  Director Score: {ss.get('dir_score',0)}  "
                               f"|  Distribution: {ss.get('dist_score',0)}  "
                               f"|  Governance Penalty: {ss.get('governance_penalty',0)}"
                               f"  |  Formula: {ss.get('formula','')}")
            sh.font      = Font(name="Arial", size=8, italic=True)
            sh.alignment = Alignment(horizontal="left", vertical="center")
            sh.border    = THIN
            row += 1

        # Dealability signals
        signals = deal.get("signals", [])
        if signals:
            ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
            sig_text = "  Dealability Signals:  " + "  |  ".join(
                f"{s['type']} ({s.get('date','')[:7]})" for s in signals[:4])
            sc = ws.cell(row=row, column=1, value=sig_text)
            sc.font      = Font(name="Arial", size=8, italic=True, color="1F3864")
            sc.alignment = Alignment(horizontal="left", vertical="center")
            sc.border    = THIN
            row += 1

        # Sell intent signals
        si = c.get("sell_intent", {})
        if si.get("sell_intent_score") is not None:
            ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
            si_sigs = "  |  ".join(si.get("sell_signals", [])[:4])
            si_txt  = (f"  Sell Intent: {si.get('sell_intent_score')} / 100  "
                       f"({si.get('sell_intent_band','')})  |  {si_sigs}")
            si_c = ws.cell(row=row, column=1, value=si_txt)
            si_c.fill      = sell_intent_fill(si.get("sell_intent_band", ""))
            si_c.font      = Font(name="Arial", size=8, italic=True,
                                  color=sell_intent_font(si.get("sell_intent_band", "")))
            si_c.alignment = Alignment(horizontal="left", vertical="center")
            si_c.border    = THIN
            row += 1

        # Directors
        dirs = c.get("directors", [])
        if dirs:
            ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
            dh = ws.cell(row=row, column=1, value="Directors / Officers  (Tier 1 — Companies House)")
            dh.fill      = fill(DKGREY)
            dh.font      = Font(name="Arial", bold=True, size=8, color=WHITE)
            dh.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            for d in dirs[:6]:
                age_val = d.get("age") or d.get("age_est")
                age_str = f"Age ~{age_val}" if age_val else "Age unknown"
                # years_active: use stored value, or derive from appointed date
                yrs_active = d.get("years_active")
                if yrs_active is None:
                    appt = (d.get("appointed") or "")[:4]
                    try:
                        yrs_active = round(2026 - int(appt), 1) if len(appt) == 4 else 0
                    except (ValueError, TypeError):
                        yrs_active = 0
                occupation = (d.get("occupation") or d.get("role") or "")[:40]
                txt = (f"  {d['name'].title()}  |  {age_str}  "
                       f"|  Appointed: {(d.get('appointed') or '')[:7]}  "
                       f"|  Tenure: {yrs_active:.1f} yrs  "
                       f"|  {occupation}")
                ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
                dc = ws.cell(row=row, column=1, value=txt)
                dc.font      = Font(name="Arial", size=8)
                dc.alignment = Alignment(horizontal="left", vertical="center")
                dc.border    = THIN
                row += 1

        ws.row_dimensions[row].height = 8
        row += 1

    for ci in range(1, n + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["N"].width = 22


# ── Sheet 3: Director contacts ────────────────────────────────────────────────

def build_contacts(wb, companies):
    ws = wb.create_sheet("Director Contacts")
    n  = 10

    title_row(ws, 1, n, "DIRECTOR CONTACT INTELLIGENCE")
    sub_row(ws, 2, n,
            "High = verified on site  |  Medium = pattern inferred from confirmed domain  "
            "|  Low = domain unconfirmed  |  Data tier shown per record")

    headers = ["Rank", "Company", "Director Name", "Role", "Age",
               "Best Email", "Confidence", "Email Pattern", "Website", "Data Tier"]
    widths  = [6, 40, 28, 20, 7, 38, 12, 30, 35, 25]

    ws.row_dimensions[3].height = 20
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 4
    for rank, c in enumerate(companies[:50], 1):
        contacts = c.get("contacts", {})
        dir_contacts = contacts.get("director_contacts", [])
        website = contacts.get("website", {})
        site_url = (website.get("website_url", "") or
                    c.get("website_url", "") or
                    c.get("digital_health", {}).get("website_url", "") or "")

        # ── Fallback: build lightweight rows from raw directors list ──────────
        if not dir_contacts:
            raw_dirs = c.get("directors", [])
            if raw_dirs:
                for di, d in enumerate(raw_dirs[:6]):
                    bg = ALT if rank % 2 == 0 else None
                    age_val = d.get("age") or d.get("age_est")
                    age_disp = str(age_val) if age_val else "-"
                    cell(ws, row, 1, rank if di == 0 else "",        bg=bg, align="center")
                    cell(ws, row, 2, c["company_name"] if di == 0 else "", bg=bg)
                    cell(ws, row, 3, (d.get("name") or "").title(),  bg=bg)
                    cell(ws, row, 4, d.get("role", ""),              bg=bg)
                    cell(ws, row, 5, age_disp,                       bg=bg, align="center")
                    cell(ws, row, 6, "",                             bg=bg)          # email — not enriched
                    cell(ws, row, 7, "—",                            bg=bg, align="center")
                    cell(ws, row, 8, "",                             bg=bg)
                    cell(ws, row, 9, site_url if di == 0 else "",    bg=bg)
                    cell(ws, row, 10, "Raw CH data — no email enrichment", bg=bg, size=8)
                    row += 1
            else:
                bg = ALT if rank % 2 == 0 else None
                cell(ws, row, 1, rank,              bg=bg, align="center")
                cell(ws, row, 2, c["company_name"], bg=bg)
                cell(ws, row, 3, "No director data available", bg=bg, fg="888888")
                cell(ws, row, 9, site_url,          bg=bg)
                row += 1
            continue

        for di, d in enumerate(dir_contacts):
            bg = ALT if rank % 2 == 0 else None
            conf = d.get("email_confidence", "None")
            conf_bg = GREEN if conf == "High" else (AMBER if conf == "Medium" else (RED if conf == "Low" else bg))
            age_val = d.get("age") or d.get("age_est")
            age_disp = str(age_val) if age_val else "-"

            cell(ws, row, 1, rank if di == 0 else "",  bg=bg, align="center")
            cell(ws, row, 2, c["company_name"] if di == 0 else "", bg=bg)
            cell(ws, row, 3, d.get("name", ""),        bg=bg)
            cell(ws, row, 4, d.get("role", ""),        bg=bg)
            cell(ws, row, 5, age_disp,                 bg=bg, align="center")
            cell(ws, row, 6, d.get("best_email", ""),  bg=conf_bg)
            cell(ws, row, 7, conf,                     bg=conf_bg, align="center", bold=True)
            patterns = d.get("email_patterns", [])
            cell(ws, row, 8, patterns[0].get("pattern","") if patterns else "", bg=bg)
            cell(ws, row, 9, site_url if di == 0 else "", bg=bg)
            cell(ws, row, 10, d.get("data_tier",""),   bg=bg, size=8)
            row += 1

    ws.freeze_panes = "A4"


# ── Sheet 4: Financials ───────────────────────────────────────────────────────

def build_financials(wb, companies):
    ws = wb.create_sheet("Financial Estimates")
    n  = 14

    title_row(ws, 1, n, "FINANCIAL ESTIMATION — PE Triangulation Revenue Model")
    sub_row(ws, 2, n,
            "6-model PE triangulation: Employee × RPE  |  Asset Turnover  |  "
            "Staff Cost reverse-engineering  |  Net Asset scaling  |  "
            "Location/Site  |  Director Salary Hybrid  —  "
            "Sector-specific benchmarks applied  |  Actual turnover not disclosed (Total Exemption filers)")

    headers = ["Rank", "Company", "Accts Type", "Period End",
               "Rev Low (£)", "Rev Base (£)", "Rev High (£)", "Confidence",
               "EBITDA Low", "EBITDA Base", "EBITDA High",
               "Net Assets", "Charges", "Formula / Source"]
    widths  = [6, 42, 18, 12, 14, 14, 14, 12, 13, 13, 13, 14, 9, 45]

    ws.row_dimensions[3].height = 20
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    def fmt(v):
        return f"£{v:,.0f}" if v else "-"

    row = 4
    for rank, c in enumerate(companies, 1):
        bg  = ALT if rank % 2 == 0 else None
        fin    = c.get("financials") or {}
        rev    = fin.get("revenue_estimate") or {}
        ebitda = fin.get("ebitda_estimate") or {}
        bs     = fin.get("balance_sheet") or c.get("bs") or {}
        ch     = c.get("charges") or fin.get("charges") or {}
        ratios = fin.get("balance_sheet_ratios") or {}

        # Confidence colouring
        conf = rev.get("confidence") or c.get("confidence", "")
        conf_bg = (fill("E2EFDA") if conf == "HIGH" else
                   fill("FFF2CC") if conf == "MEDIUM" else
                   fill("FFD6D6") if conf == "LOW" else bg)

        # EBITDA — use estimate dict first, then fall back to top-level scalar
        eb_low  = ebitda.get("ebitda_low")  or c.get("ebitda_low")
        eb_base = ebitda.get("ebitda_base") or c.get("ebitda_base")
        eb_high = ebitda.get("ebitda_high") or c.get("ebitda_high")

        cell(ws, row, 1,  rank,                                  bg=bg, align="center")
        cell(ws, row, 2,  c["company_name"],                     bg=bg)
        cell(ws, row, 3,  bs.get("accounts_type") or "N/A",     bg=bg, size=8)
        cell(ws, row, 4,  (bs.get("period_end") or "")[:7],     bg=bg, align="center")
        cell(ws, row, 5,  fmt(rev.get("revenue_low")  or c.get("rev_low")),  bg=bg, align="right")
        cell(ws, row, 6,  fmt(rev.get("revenue_base") or c.get("rev_base")), bg=bg, align="right", bold=True)
        cell(ws, row, 7,  fmt(rev.get("revenue_high") or c.get("rev_high")), bg=bg, align="right")
        cell(ws, row, 8,  conf,                                  bg=conf_bg, align="center", bold=True)
        cell(ws, row, 9,  fmt(eb_low),                           bg=bg, align="right")
        cell(ws, row, 10, fmt(eb_base),                          bg=bg, align="right", bold=True)
        cell(ws, row, 11, fmt(eb_high),                          bg=bg, align="right")
        cell(ws, row, 12, fmt(ratios.get("net_assets") or bs.get("net_assets")), bg=bg, align="right")
        cell(ws, row, 13, ch.get("outstanding_charges") if ch.get("outstanding_charges") is not None else "-",
             bg=bg, align="center")
        formula_str = rev.get("formula") or (
            "Tier 4 — " + ", ".join(c.get("models_used", [])) if c.get("models_used") else "")
        cell(ws, row, 14, formula_str, bg=bg, size=8)
        row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(companies)+3}"


# ── Sheet 5: Bolt-on analysis ─────────────────────────────────────────────────

def build_bolt_on(wb, bolt_on_data: dict):
    ws = wb.create_sheet("Bolt-On Analysis")

    title_row(ws, 1, 8, "BOLT-ON OPPORTUNITY ANALYSIS — Sector Adjacency & Roll-Up Map")

    # Market fragmentation
    row = 3
    frag = bolt_on_data.get("market_fragmentation", {})
    ws.merge_cells(f"A{row}:H{row}")
    fr = ws.cell(row=row, column=1,
                 value=f"  Market Fragmentation Index: {frag.get('fragmentation_index','-')}  |  "
                       f"{frag.get('interpretation','')}  |  "
                       f"Formula: {frag.get('formula','')}  |  "
                       f"Total Companies Analysed: {frag.get('total_companies_analysed','-')}")
    fr.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    fr.fill      = fill(NAVY)
    fr.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 2

    # Bolt-on recommendations
    for rec in bolt_on_data.get("bolt_on_recommendations", []):
        ws.merge_cells(f"A{row}:H{row}")
        rh = ws.cell(row=row, column=1,
                     value=f"  {rec['cluster']}  —  Opportunity Score: {rec['opportunity_score']}/10")
        rh.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        rh.fill      = fill(BLUE)
        rh.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        cell(ws, row, 1, "Rationale",      bg=GREY, bold=True, size=9)
        ws.merge_cells(f"B{row}:H{row}")
        cell(ws, row, 2, rec["rationale"],  bg=None, size=9)
        row += 1

        cell(ws, row, 1, "Bolt-On Services", bg=GREY, bold=True, size=9)
        ws.merge_cells(f"B{row}:H{row}")
        cell(ws, row, 2, "  •  ".join(rec["bolt_on_services"]), bg=None, size=9)
        row += 1

        cell(ws, row, 1, "Example Targets", bg=GREY, bold=True, size=9)
        targets_text = "  |  ".join(
            f"{t['name']} (Score: {t['score']})" for t in rec.get("example_targets", [])[:5]
        )
        ws.merge_cells(f"B{row}:H{row}")
        cell(ws, row, 2, targets_text or "—", bg=ALT, size=9)
        row += 1

        ws.row_dimensions[row].height = 8
        row += 1

    # SIC clusters
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    sc = ws.cell(row=row, column=1, value="SIC CODE DISTRIBUTION")
    sc.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
    sc.fill  = fill(DKGREY)
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    for cluster in bolt_on_data.get("sic_clusters", []):
        cell(ws, row, 1, cluster["sic_code"],          bg=GREY, bold=True, align="center")
        cell(ws, row, 2, f"{cluster['company_count']} companies", bg=GREY)
        ws.merge_cells(f"C{row}:H{row}")
        cell(ws, row, 3, ", ".join(cluster["example_companies"][:5]), bg=None, size=8)
        row += 1

    for ci, w in enumerate([14, 45, 45, 1, 1, 1, 1, 1], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w if w > 1 else 5


# ── Sheet 6: Sell Signals ─────────────────────────────────────────────────────

def build_sell_signals(wb, companies):
    ws = wb.create_sheet("Sell Signals")
    n  = 12

    title_row(ws, 1, n, "SELL INTENT ANALYSIS — Owner Exit Readiness Signals")
    sub_row(ws, 2, n,
            "Score 0–100: Strong(70+) = priority outreach  |  Moderate(50–69) = pipeline  |  "
            "Dimensions: Age/Tenure(40) + Structure(25) + Stress(20) + Maturity(15)  |  Tier 1 — Companies House")

    headers = ["Rank", "Company", "Sell Intent Score", "Band",
               "Age/Tenure", "Structure", "Stress", "Maturity",
               "Late Filings", "Dir. Departs", "Sell Signals Summary"]
    widths  = [6, 44, 14, 12, 12, 12, 10, 10, 12, 12, 70]

    ws.row_dimensions[3].height = 22
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for rank, c in enumerate(companies, 1):
        row = rank + 3
        bg  = ALT if rank % 2 == 0 else None
        si  = c.get("sell_intent", {})
        if not si:
            cell(ws, row, 1, rank, bg=bg, align="center")
            cell(ws, row, 2, c["company_name"], bg=bg)
            for col in range(3, n + 1):
                cell(ws, row, col, "-", bg=bg, align="center")
            continue

        band     = si.get("sell_intent_band", "")
        score    = si.get("sell_intent_score", 0)
        comps    = si.get("components", {})
        at_sc    = comps.get("age_tenure", {}).get("score", 0)
        st_sc    = comps.get("business_structure", {}).get("score", 0)
        op_sc    = comps.get("operational_stress", {}).get("score", 0)
        mat_sc   = comps.get("company_maturity", {}).get("score", 0)
        late     = comps.get("operational_stress", {}).get("late_filings", 0)
        departs  = comps.get("operational_stress", {}).get("resignations_3yr", 0)
        sigs     = "  |  ".join(si.get("sell_signals", [])[:5])

        cell(ws, row, 1, rank,  bg=bg, align="center", bold=True)
        cell(ws, row, 2, c["company_name"], bg=bg)

        for col, val in [(3, score), (4, band)]:
            cx = ws.cell(row=row, column=col, value=val)
            cx.fill      = sell_intent_fill(band)
            cx.font      = Font(name="Arial", size=9, bold=True, color=sell_intent_font(band))
            cx.alignment = Alignment(horizontal="center", vertical="center")
            cx.border    = THIN

        cell(ws, row, 5,  at_sc,   bg=bg, align="center")
        cell(ws, row, 6,  st_sc,   bg=bg, align="center")
        cell(ws, row, 7,  op_sc,   bg=bg, align="center")
        cell(ws, row, 8,  mat_sc,  bg=bg, align="center")
        cell(ws, row, 9,  late,    bg=RED if late > 0 else bg,    align="center")
        cell(ws, row, 10, departs, bg=AMBER if departs > 0 else bg, align="center")
        cell(ws, row, 11, sigs,    bg=bg, size=8, wrap=True)
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(companies)+3}"


# ── Sheet 7: Government Contracts ─────────────────────────────────────────────

def build_contracts(wb, companies):
    ws = wb.create_sheet("Gov. Contracts")
    n  = 10

    title_row(ws, 1, n, "GOVERNMENT CONTRACT INTELLIGENCE — Revenue Quality Signals")
    sub_row(ws, 2, n,
            "Source: Contracts Finder + Find a Tender  |  Tier 1 — Public procurement registers  |  "
            "Government revenue = recurring, creditworthy counterparty — multiple = strong PE valuation signal")

    headers = ["Rank", "Company", "Contracts Found", "Total Value (£)",
               "Latest Date", "Buyers", "Revenue Quality", "Top Contract Title"]
    widths  = [6, 44, 15, 16, 14, 45, 42, 60]

    ws.row_dimensions[3].height = 22
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Filter to companies with government contract data
    with_contracts = [c for c in companies if c.get("government_contracts", {}).get("contracts_found", 0) > 0]
    no_contracts   = [c for c in companies if c.get("government_contracts", {}).get("contracts_found", 0) == 0
                      and c.get("government_contracts", {}).get("revenue_quality", "") != "Not searched"]

    row = 4
    # Show companies WITH contracts first
    ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
    hdr = ws.cell(row=row, column=1, value=f"  COMPANIES WITH GOVERNMENT CONTRACTS ({len(with_contracts)} found)")
    hdr.fill = fill(TEAL); hdr.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    for rank, c in enumerate(with_contracts, 1):
        bg  = ALT if rank % 2 == 0 else None
        gc  = c.get("government_contracts", {})
        cl  = gc.get("contract_list", [])
        top = cl[0].get("title", "") if cl else ""
        buyers_str = "  |  ".join(gc.get("buyers", [])[:3])

        cell(ws, row, 1, rank,                           bg=bg, align="center", bold=True)
        cell(ws, row, 2, c["company_name"],              bg=bg)
        cx3 = ws.cell(row=row, column=3, value=gc.get("contracts_found", 0))
        cx3.fill = fill(GREEN); cx3.font = Font(name="Arial", size=9, bold=True)
        cx3.alignment = Alignment(horizontal="center", vertical="center"); cx3.border = THIN
        tv = gc.get("total_contract_value", 0)
        cell(ws, row, 4, f"£{tv:,.0f}" if tv else "-",  bg=bg, align="right", bold=True)
        cell(ws, row, 5, gc.get("latest_contract_date","")[:10], bg=bg, align="center")
        cell(ws, row, 6, buyers_str,                     bg=bg, size=8)
        cell(ws, row, 7, gc.get("revenue_quality",""),   bg=GREEN, size=8)
        cell(ws, row, 8, top[:80],                       bg=bg, size=8)
        row += 1

    # Summary of no-contract companies
    if no_contracts:
        row += 1
        ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
        nh = ws.cell(row=row, column=1, value=f"  NO GOVERNMENT CONTRACTS DETECTED ({len(no_contracts)} companies searched)")
        nh.fill = fill(GREY); nh.font = Font(name="Arial", bold=True, size=9)
        nh.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        for rank, c in enumerate(no_contracts[:20], 1):
            bg = ALT if rank % 2 == 0 else None
            cell(ws, row, 1, rank,              bg=bg, align="center")
            cell(ws, row, 2, c["company_name"], bg=bg)
            for col in range(3, n + 1):
                cell(ws, row, col, "-", bg=bg, align="center")
            row += 1

    ws.freeze_panes = "A4"


# ── Sheet 8: Digital Health ───────────────────────────────────────────────────

def build_digital_health(wb, companies):
    ws = wb.create_sheet("Digital Health")
    n  = 14

    title_row(ws, 1, n, "DIGITAL HEALTH & SECTOR VERIFICATION — Online Presence, Maturity & Sector Confirmation")
    sub_row(ws, 2, n,
            "Digital Score 0–100: Mature(80+) | Adequate(60–79) | Below Average(40–59) | Poor(<40)  |  "
            "Sector Match: Confirmed(70+) | Likely(40–69) | Uncertain(<40) | Unverified(no website)  |  Tier 3 — Website analysis + WHOIS")

    headers = ["Rank", "Company", "Digital Score", "Band", "Domain Age (yrs)",
               "Website Live", "LinkedIn", "Job Postings",
               "Accreditations Detected", "Accred. Score", "Domain",
               "Sector Match", "Sector Match Score", "Match Signals"]
    widths  = [6, 44, 13, 14, 15, 12, 10, 12, 50, 13, 35, 14, 14, 60]

    ws.row_dimensions[3].height = 22
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    assessed = [c for c in companies if c.get("digital_health", {}).get("digital_health_band", "Not assessed") != "Not assessed"]

    for rank, c in enumerate(assessed, 1):
        row = rank + 3
        bg  = ALT if rank % 2 == 0 else None
        dh  = c.get("digital_health", {})
        ac  = c.get("accreditations", {})
        ds  = dh.get("digital_health_score")
        band= dh.get("digital_health_band", "")

        cell(ws, row, 1, rank, bg=bg, align="center", bold=True)
        cell(ws, row, 2, c["company_name"], bg=bg)

        # Digital score + band (colour-coded)
        for col, val in [(3, ds), (4, band)]:
            cx = ws.cell(row=row, column=col, value=val if val is not None else "-")
            cx.fill      = digital_band_fill(band) if band else fill(GREY)
            cx.font      = Font(name="Arial", size=9, bold=True,
                               color=WHITE if band in ("Mature",) else "000000")
            cx.alignment = Alignment(horizontal="center", vertical="center")
            cx.border    = THIN

        da = dh.get("domain_age_years")
        cell(ws, row, 5, f"{da:.1f}" if da else "Unknown",  bg=bg, align="center")

        live = dh.get("website_live", False)
        cell(ws, row, 6, "✓" if live else "✗",
             bg=GREEN if live else RED, align="center", bold=True)

        li = dh.get("has_linkedin", False)
        cell(ws, row, 7, "✓" if li else "✗",
             bg=GREEN if li else bg, align="center")

        jobs = dh.get("has_job_postings", False)
        cell(ws, row, 8, "Hiring" if jobs else "-",
             bg=GREEN if jobs else bg, align="center")

        accreds_site = dh.get("accreditations_on_site", [])
        accreds_all  = ac.get("accreditations", [])
        all_accreds  = list(set(accreds_site + accreds_all))
        cell(ws, row, 9,  "  |  ".join(all_accreds[:6]) or "-",    bg=bg, size=8)
        cell(ws, row, 10, ac.get("accreditation_score", "-"),      bg=bg, align="center")
        cell(ws, row, 11, dh.get("domain", ""),                    bg=bg, size=8)

        # ── Sector relevance columns ──────────────────────────────────────────
        srl   = dh.get("sector_relevance_label", "Unverified")
        srs   = dh.get("sector_relevance_score", 0)
        srsig = " | ".join(dh.get("sector_match_signals", []))

        # Colour-code the sector match label
        match_fill = {
            "Confirmed":  fill(GREEN),
            "Likely":     fill("FFF2CC"),   # soft yellow
            "Uncertain":  fill("FFD6D6"),   # light red
            "Unverified": fill("EEEEEE"),   # grey
        }.get(srl, fill("EEEEEE"))

        cx12 = ws.cell(row=row, column=12, value=srl)
        cx12.fill      = match_fill
        cx12.font      = Font(name="Arial", size=9, bold=True)
        cx12.alignment = Alignment(horizontal="center", vertical="center")
        cx12.border    = THIN

        cell(ws, row, 13, srs,   bg=bg, align="center")
        cell(ws, row, 14, srsig, bg=bg, size=8)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(assessed)+3}"


# ── Sheet 9: Regulatory Registers ────────────────────────────────────────────

REG_DISPLAY = {
    "EA_WASTE":    ("EA Waste Permit",    "EA"),
    "EA_CARRIERS": ("EA Carrier/Broker",  "EA"),
    "CQC":         ("CQC Registered",     "CQC"),
    "FCA":         ("FCA Authorised",     "FCA"),
    "ICO":         ("ICO Data Controller","ICO"),
    "OFSTED":      ("Ofsted Provider",    "Ofsted"),
    "SIA":         ("SIA Approved",       "SIA"),
}


def build_regulatory(wb, companies):
    ws = wb.create_sheet("Regulatory Registers")

    reg_keys = list(REG_DISPLAY.keys())
    n_fixed  = 5   # Rank, Company, Reg Score, Band, Confirmed
    n_cols   = n_fixed + len(reg_keys) + 2  # + Accred Score + Combined

    title_row(ws, 1, n_cols,
              "REGULATORY REGISTER VERIFICATION — UK Public Register Checks")
    sub_row(ws, 2, n_cols,
            "Tier 1 public registers: EA Waste | EA Carriers | CQC | FCA | ICO | Ofsted | SIA  |  "
            "✅ = Confirmed registration  ❌ = Not found / N/A  |  Score 0–25")

    col_headers = (
        ["Rank", "Company", "Reg. Score", "Reg. Band", "Confirmed Registrations"] +
        [REG_DISPLAY[k][0] for k in reg_keys] +
        ["Accred. Score", "Combined Score"]
    )
    col_widths = (
        [6, 42, 11, 18, 60] +
        [16] * len(reg_keys) +
        [12, 13]
    )

    ws.row_dimensions[3].height = 28
    for ci, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell(ws, 3, ci, h, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Only include companies that had accreditation enrichment run
    assessed = [
        c for c in companies
        if c.get("accreditations", {}).get("combined_band", "Not assessed") != "Not assessed"
    ]
    # Sort by combined score descending
    assessed.sort(
        key=lambda c: c.get("accreditations", {}).get("combined_score", 0) or 0,
        reverse=True,
    )

    for rank, c in enumerate(assessed, 1):
        row = rank + 3
        bg  = ALT if rank % 2 == 0 else None
        ac  = c.get("accreditations", {})
        regs = ac.get("registrations", {})

        reg_score = ac.get("regulatory_score")
        reg_band  = ac.get("regulatory_band", "")
        confirmed = "  |  ".join(ac.get("confirmed_regs", []))[:80] or "—"
        accred_s  = ac.get("accreditation_score")
        combined  = ac.get("combined_score")
        comb_band = ac.get("combined_band", "")

        cell(ws, row, 1, rank, bg=bg, align="center", bold=True)
        cell(ws, row, 2, c["company_name"], bg=bg)

        # Regulatory score — colour by band
        rs_bg = fill("375623") if (reg_score or 0) >= 15 else \
                fill("70AD47") if (reg_score or 0) >= 9  else \
                fill(AMBER)    if (reg_score or 0) >= 4  else fill(GREY)
        cx = ws.cell(row=row, column=3, value=reg_score if reg_score is not None else "—")
        cx.fill = rs_bg; cx.font = Font(name="Arial", size=9, bold=True, color="000000")
        cx.alignment = Alignment(horizontal="center", vertical="center"); cx.border = THIN

        cell(ws, row, 4, reg_band or "—", bg=bg, size=9)
        cell(ws, row, 5, confirmed, bg=bg, size=8, wrap=True)

        # Individual register columns
        for ci, key in enumerate(reg_keys, n_fixed + 1):
            r = regs.get(key, {})
            found = r.get("found", False) if isinstance(r, dict) else False
            ref   = (r.get("permit_reference") or r.get("cqc_provider_id") or
                     r.get("fca_ref") or r.get("ico_reg_number") or "") if isinstance(r, dict) else ""
            label = f"✅ {ref[:12]}" if (found and ref) else "✅" if found else "❌"
            cell(ws, row, ci, label,
                 bg=GREEN if found else (RED if isinstance(r, dict) and "found" in r else bg),
                 align="center", size=9)

        cell(ws, row, n_fixed + len(reg_keys) + 1,
             accred_s if accred_s is not None else "—", bg=bg, align="center")

        # Combined score
        comb_bg = fill("375623") if (combined or 0) >= 35 else \
                  fill("70AD47") if (combined or 0) >= 25 else \
                  fill(AMBER)    if (combined or 0) >= 15 else fill(GREY)
        cx2 = ws.cell(row=row, column=n_fixed + len(reg_keys) + 2,
                      value=combined if combined is not None else "—")
        cx2.fill = comb_bg; cx2.font = Font(name="Arial", size=9, bold=True, color="000000")
        cx2.alignment = Alignment(horizontal="center", vertical="center"); cx2.border = THIN

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{len(assessed)+3}"


# ── Sheet 10: Summary stats ───────────────────────────────────────────────────

def build_summary(wb, companies):
    ws = wb.create_sheet("Summary Stats")
    title_row(ws, 1, 4, "PIPELINE SUMMARY STATISTICS")

    acq_scores = [c["acquisition_score"] for c in companies]

    si_scores = [c.get("sell_intent",{}).get("sell_intent_score",0) or 0 for c in companies]

    stats = [
        ("PIPELINE OVERVIEW", ""),
        ("Total companies found",           len(companies)),
        ("Prime targets  (score ≥ 80)",     sum(1 for s in acq_scores if s >= 80)),
        ("High priority  (65–79)",          sum(1 for s in acq_scores if 65 <= s < 80)),
        ("Medium priority  (50–64)",        sum(1 for s in acq_scores if 50 <= s < 65)),
        ("Intelligence only  (< 50)",       sum(1 for s in acq_scores if s < 50)),
        ("", ""),
        ("SELL INTENT SIGNALS", ""),
        ("Strong sell intent  (score 70+)",    sum(1 for c in companies if (c.get("sell_intent",{}).get("sell_intent_score") or 0) >= 70)),
        ("Moderate sell intent  (50–69)",      sum(1 for c in companies if 50 <= (c.get("sell_intent",{}).get("sell_intent_score") or 0) < 70)),
        ("Weak sell intent  (30–49)",          sum(1 for c in companies if 30 <= (c.get("sell_intent",{}).get("sell_intent_score") or 0) < 50)),
        ("Low sell intent  (< 30)",            sum(1 for c in companies if 0 < (c.get("sell_intent",{}).get("sell_intent_score") or 0) < 30)),
        ("Late filing history detected",       sum(1 for c in companies if (c.get("sell_intent",{}).get("components",{}).get("operational_stress",{}).get("late_filings",0)) > 0)),
        ("Director departure in last 3 yrs",   sum(1 for c in companies if (c.get("sell_intent",{}).get("components",{}).get("operational_stress",{}).get("resignations_3yr",0)) > 0)),
        ("", ""),
        ("OWNERSHIP & SUCCESSION", ""),
        ("PE-backed (flagged for exclusion)",  sum(1 for c in companies if c.get("pe_backed"))),
        ("Family / owner-managed",             sum(1 for c in companies if c.get("is_family"))),
        ("Solo director (key-person risk)",    sum(1 for c in companies if c.get("director_count") == 1)),
        ("Max director age 65+",               sum(1 for c in companies if (c.get("succession") or {}).get("max_age",0) >= 65)),
        ("Max director age 55+",               sum(1 for c in companies if (c.get("succession") or {}).get("max_age",0) >= 55)),
        ("Succession score ≥ 80",              sum(1 for c in companies if (c.get("succession") or {}).get("total",0) >= 80)),
        ("", ""),
        ("COMPANY AGE", ""),
        ("Incorporated pre-2000  (25+ yrs)",   sum(1 for c in companies if c.get("company_age_years",0) >= 25)),
        ("Incorporated 2000–2009",             sum(1 for c in companies if 15 <= c.get("company_age_years",0) < 25)),
        ("Incorporated 2010–2019",             sum(1 for c in companies if 5 <= c.get("company_age_years",0) < 15)),
        ("Incorporated 2020+",                 sum(1 for c in companies if c.get("company_age_years",0) < 5)),
        ("", ""),
        ("DEALABILITY", ""),
        ("With dealability signals",           sum(1 for c in companies if c.get("dealability",{}).get("signal_count",0) > 0)),
        ("With outstanding charges (debt)",    sum(1 for c in companies if c.get("charges",{}).get("outstanding_charges",0) > 0)),
        ("Clean charge register",              sum(1 for c in companies if c.get("charges",{}).get("outstanding_charges",0) == 0)),
        ("", ""),
        ("DIGITAL & ACCREDITATIONS", ""),
        ("Live website confirmed",             sum(1 for c in companies if c.get("digital_health",{}).get("website_live",False))),
        ("LinkedIn presence detected",         sum(1 for c in companies if c.get("digital_health",{}).get("has_linkedin",False))),
        ("Active job postings detected",       sum(1 for c in companies if c.get("digital_health",{}).get("has_job_postings",False))),
        ("Government contracts found",         sum(1 for c in companies if (c.get("government_contracts",{}).get("contracts_found",0) or 0) > 0)),
        ("With detected accreditations",       sum(1 for c in companies if (c.get("accreditations",{}).get("accreditation_count",0) or 0) > 0)),
        ("", ""),
        ("REGULATORY REGISTERS", ""),
        ("Any register confirmed",             sum(1 for c in companies if (c.get("accreditations",{}).get("reg_count",0) or 0) > 0)),
        ("EA Waste permit holders",            sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("EA_WASTE",{}).get("found",False))),
        ("EA Carrier / Broker registered",     sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("EA_CARRIERS",{}).get("found",False))),
        ("CQC registered provider",            sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("CQC",{}).get("found",False))),
        ("FCA authorised firm",                sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("FCA",{}).get("found",False))),
        ("ICO data controller",                sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("ICO",{}).get("found",False))),
        ("Ofsted registered provider",         sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("OFSTED",{}).get("found",False))),
        ("SIA Approved Contractor",            sum(1 for c in companies if c.get("accreditations",{}).get("registrations",{}).get("SIA",{}).get("found",False))),
        ("Regulatory score ≥ 15 (Highly Reg.)",sum(1 for c in companies if (c.get("accreditations",{}).get("regulatory_score") or 0) >= 15)),
        ("Combined score ≥ 25 (Strong)",       sum(1 for c in companies if (c.get("accreditations",{}).get("combined_score") or 0) >= 25)),
    ]

    for ri, (label, val) in enumerate(stats, 2):
        ws.row_dimensions[ri].height = 18
        if val == "":   # section header
            cell(ws, ri, 1, label, bg=NAVY, fg=WHITE, bold=True, size=10, align="left")
            ws.merge_cells(f"A{ri}:B{ri}")
        else:
            bg = ALT if ri % 2 == 0 else None
            cell(ws, ri, 1, label, bg=bg, bold=False, size=10)
            cell(ws, ri, 2, val,   bg=bg, bold=True, fg=NAVY, align="center", size=10)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12


# ── Sheet 11: Competitor Map ──────────────────────────────────────────────────

def _load_sic_descriptions() -> dict[str, str]:
    """Load SIC code → human-readable description from data/sic_codes.json."""
    sic_path = os.path.join(os.path.dirname(__file__), "data", "sic_codes.json")
    try:
        with open(sic_path) as f:
            raw = json.load(f)
        # Values may be dicts {"description": "...", "count": N} or plain strings
        return {
            k: (v["description"] if isinstance(v, dict) else str(v))
            for k, v in raw.items()
        }
    except Exception:
        return {}


def _sic_labels(sic_codes: list[str], sic_desc: dict[str, str]) -> list[str]:
    """Translate a list of SIC codes to short description labels."""
    labels = []
    for code in sic_codes:
        desc = sic_desc.get(str(code), "")
        # Truncate long descriptions to keep cells readable
        if desc:
            labels.append(desc[:70] + ("…" if len(desc) > 70 else ""))
        else:
            labels.append(code)  # fallback to raw code
    return labels


def build_competitors(wb, companies):
    """
    One row per competitor per target company.
    Columns include:
      - Distance in miles + band (colour-coded)
      - Competitor services (SIC descriptions)
      - Services not offered by the target (SIC gap — competitor has, target lacks)
      - Revenue, PE-backed, group, acquisition fit, sell intent
    """
    ws = wb.create_sheet("Competitor Map")
    headers = [
        ("Rank",                   5),
        ("Target Company",        38),
        ("#",                      4),
        ("Competitor Name",       38),
        ("Reg. No.",              12),
        ("Postcode",              10),
        ("Distance (mi)",         13),
        ("Band",                  15),
        ("Est. Revenue £",        14),
        ("Accts Type",            11),
        ("PE-backed",              9),
        ("Group",                  7),
        ("Acq. Fit",               9),
        ("Sell Intent",           10),
        ("Website",               32),
        ("Competitor Services",   52),
        ("Services Not in Target", 52),
    ]
    n = len(headers)

    title_row(ws, 1, n, "COMPETITOR MAP — Nearest competitors per target company  (sorted by distance)")
    sub_row(ws, 2, n,
            "Distance = haversine miles between registered postcodes  ·  "
            "Local ≤15mi  ·  Regional 15–50mi  ·  Adjacent Region 50–100mi  ·  National >100mi  ·  "
            "Red = PE-backed/group  ·  Orange = services the target does NOT offer")

    ws.row_dimensions[3].height = 24
    for ci, (label, width) in enumerate(headers, 1):
        cell(ws, 3, ci, label, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Load SIC descriptions once
    sic_desc = _load_sic_descriptions()

    # Band and fit colour maps (defined once, reused per row)
    band_fill_map = {
        "Local":           fill("E2EFDA"),
        "Regional":        fill("FFF2CC"),
        "Adjacent Region": fill("FFE0B2"),
        "National":        fill("D9D9D9"),
    }
    fit_fill_map = {
        "High":   fill("E2EFDA"),
        "Medium": fill("FFF2CC"),
        "Low":    fill("FFD6D6"),
    }

    row = 4
    for rank, c in enumerate(companies, 1):
        comp_analysis = c.get("competitor_analysis") or {}
        comp_map      = comp_analysis.get("competitor_map", [])

        # Target's own SIC codes (normalised to strings)
        target_sics = set(str(s) for s in (c.get("sic_codes") or []) if s)

        if not comp_map:
            bg = ALT if rank % 2 == 0 else None
            cell(ws, row, 1, rank, bg=bg, align="center")
            cell(ws, row, 2, c["company_name"], bg=bg)
            cell(ws, row, 3, "-", bg=bg, align="center")
            cell(ws, row, 4, "No competitor data", bg=bg, fg="888888")
            for ci in range(5, n + 1):
                cell(ws, row, ci, "", bg=bg)
            row += 1
            continue

        for comp_idx, comp in enumerate(comp_map, 1):
            bg = ALT if rank % 2 == 0 else None

            is_pe  = comp.get("is_pe_backed", False)
            is_grp = comp.get("is_group_owned", False)
            row_bg = fill("FFD6D6") if (is_pe or is_grp) else bg

            miles     = comp.get("distance_miles")
            band      = comp.get("distance_band", "")
            acq_fit   = comp.get("acquisition_fit", "")
            rev       = comp.get("estimated_revenue_gbp", 0)
            si_score  = comp.get("sell_intent_score")

            # ── Service columns ──────────────────────────────────────────────
            comp_sics = [str(s) for s in (comp.get("sic_codes") or []) if s]

            # All competitor services (labelled)
            comp_services = _sic_labels(comp_sics, sic_desc)

            # Services competitor offers that target does NOT (the gap)
            gap_sics   = [s for s in comp_sics if s not in target_sics]
            gap_labels = _sic_labels(gap_sics, sic_desc)

            # ── Row cells ────────────────────────────────────────────────────
            cell(ws, row, 1,  rank, bg=bg, align="center")
            cell(ws, row, 2,  c["company_name"] if comp_idx == 1 else "",
                              bg=bg, size=8)
            cell(ws, row, 3,  comp_idx, bg=bg, align="center", size=8)
            cell(ws, row, 4,  comp.get("company_name", ""),
                              bg=row_bg, bold=is_pe or is_grp, size=8)
            cell(ws, row, 5,  comp.get("company_number", ""), bg=row_bg, size=8)
            cell(ws, row, 6,  comp.get("postcode", ""),       bg=row_bg, size=8)

            # Distance
            dist_val = f"{miles:.1f}" if miles is not None else "N/A"
            cx_dist = ws.cell(row=row, column=7, value=dist_val)
            cx_dist.fill      = band_fill_map.get(band, fill("D9D9D9"))
            cx_dist.font      = Font(name="Arial", size=9, bold=True)
            cx_dist.alignment = Alignment(horizontal="center", vertical="center")
            cx_dist.border    = THIN

            # Band
            cx_band = ws.cell(row=row, column=8, value=band)
            cx_band.fill      = band_fill_map.get(band, fill("D9D9D9"))
            cx_band.font      = Font(name="Arial", size=9)
            cx_band.alignment = Alignment(horizontal="center", vertical="center")
            cx_band.border    = THIN

            # Revenue / ownership / fit
            rev_str = f"£{rev:,.0f}" if rev else "-"
            cell(ws, row, 9,  rev_str,  bg=row_bg, align="right",  size=8)
            cell(ws, row, 10, comp.get("accounts_type", ""),
                              bg=row_bg, align="center", size=8)
            cell(ws, row, 11, "⚠ PE" if is_pe else ("Grp" if is_grp else "-"),
                              bg=fill("FFD6D6") if is_pe else row_bg,
                              align="center", size=8)
            cell(ws, row, 12, "✓" if is_grp else "-", bg=row_bg, align="center", size=8)

            cx_fit = ws.cell(row=row, column=13, value=acq_fit)
            cx_fit.fill      = fit_fill_map.get(acq_fit, fill("D9D9D9"))
            cx_fit.font      = Font(name="Arial", size=9, bold=(acq_fit == "High"))
            cx_fit.alignment = Alignment(horizontal="center", vertical="center")
            cx_fit.border    = THIN

            cell(ws, row, 14, si_score if si_score is not None else "-",
                              bg=row_bg, align="center", size=8)

            # Website URL (col 15)
            website_url = comp.get("website_url", "")
            if website_url:
                cx_web = ws.cell(row=row, column=15, value=website_url)
                cx_web.font      = Font(name="Arial", size=8, color="0563C1", underline="single")
                cx_web.alignment = Alignment(horizontal="left", vertical="center")
                cx_web.border    = THIN
                if row_bg:
                    if row_bg: cx_web.fill = row_bg if isinstance(row_bg, PatternFill) else fill(row_bg)
            else:
                cell(ws, row, 15, "—", bg=row_bg, fg="888888", align="center", size=8)

            # Competitor services (col 16) — prefer web-sourced description, fall back to SIC labels
            web_desc = comp.get("services_description", "")
            if web_desc and len(web_desc) >= 20:
                services_text = web_desc
            elif comp_services:
                services_text = "  |  ".join(comp_services)
            else:
                services_text = "-"
            cell(ws, row, 16, services_text, bg=row_bg, size=8, wrap=True)

            # Gap services (col 17) — what competitor does that target does NOT
            if gap_labels:
                gap_text = "  |  ".join(gap_labels)
                cx_gap = ws.cell(row=row, column=17, value=gap_text)
                cx_gap.fill      = fill("FFE0B2")   # orange — highlights opportunity/threat
                cx_gap.font      = Font(name="Arial", size=8, bold=True, color="7B3F00")
                cx_gap.alignment = Alignment(horizontal="left", vertical="center",
                                             wrap_text=True)
                cx_gap.border    = THIN
            else:
                cell(ws, row, 17, "—  same services", bg=row_bg, fg="888888",
                     size=8, align="center")

            # Auto-height for wrapped rows
            ws.row_dimensions[row].height = max(
                15,
                15 * max(1, len(gap_labels)),
            )

            row += 1

        # Thin spacer between targets
        ws.row_dimensions[row].height = 4
        row += 1

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}3"


# ── Sheet: Company Overviews ──────────────────────────────────────────────────

def build_overview(wb, companies):
    """
    One row per company — rich overview card pulled from their website and CH data.

    Columns:
      Rank | Company | Website | Grade | Rev. Base | Employees | About (web description)
      Services | Sector Match | Domain Age | LinkedIn | Job Postings
      Address | Incorporated | Directors | CH Link
    """
    ws = wb.create_sheet("Company Overviews")

    headers = [
        ("Rank",          5),
        ("Company",      42),
        ("Website",      30),
        ("Grade",        10),
        ("Rev. Est.",    14),
        ("Employees",    11),
        ("About",        70),   # web meta description / og:description
        ("Services",     55),   # SIC descriptions
        ("Sector Match", 14),
        ("Domain Age",   11),
        ("LinkedIn",      9),
        ("Jobs",          7),
        ("Town",         16),
        ("County",       16),
        ("Incorporated", 12),
        ("Directors",    34),
        ("CH Link",      18),
    ]
    n = len(headers)

    title_row(ws, 1, n, "COMPANY OVERVIEWS — Website-sourced intelligence per target company")
    sub_row(ws, 2, n,
            "About = homepage meta description · Services = SIC descriptions · "
            "Sector Match = web keyword verification · Source: digital_health.py + Companies House")

    ws.row_dimensions[3].height = 28
    for ci, (label, width) in enumerate(headers, 1):
        cell(ws, 3, ci, label, bg=NAVY, fg=WHITE, bold=True, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    sic_desc = _load_sic_descriptions()

    sector_fill_map = {
        "Confirmed":  fill("E2EFDA"),
        "Likely":     fill("FFF2CC"),
        "Uncertain":  fill("FFD6D6"),
        "Unverified": fill("EEEEEE"),
    }

    for i, c in enumerate(companies, 1):
        row = i + 3
        bg  = ALT if i % 2 == 0 else None

        dh       = c.get("digital_health") or {}
        contacts = c.get("contacts") or {}
        acq      = c.get("acquisition_score", 0)
        grade    = c.get("acquisition_grade", "")

        # ── Website / domain ──────────────────────────────────────────────────
        domain   = dh.get("domain", "")
        website  = contacts.get("website") or c.get("website") or (
            f"https://{domain}" if domain else "")
        website_live = dh.get("website_live", False)

        # ── About text — prefer web description, fall back to SIC desc ────────
        about = dh.get("website_description", "").strip()
        if not about:
            sic_labels = _sic_labels(
                [str(s) for s in (c.get("sic_codes") or []) if s], sic_desc)
            about = "  |  ".join(sic_labels) if sic_labels else ""

        # ── Services (SIC descriptions) ───────────────────────────────────────
        sic_labels = _sic_labels(
            [str(s) for s in (c.get("sic_codes") or []) if s], sic_desc)
        services_text = "  |  ".join(sic_labels) if sic_labels else "—"

        # ── Revenue + employees ───────────────────────────────────────────────
        _fin_ov   = c.get("financials") or {}
        _rev_ov   = _fin_ov.get("revenue_estimate") or {}
        rev_base  = _rev_ov.get("revenue_base") or c.get("rev_base")
        rev_str   = f"£{rev_base:,.0f}" if rev_base else "—"
        emp       = c.get("estimated_employees")

        # ── Sector match ──────────────────────────────────────────────────────
        srl = dh.get("sector_relevance_label", "Unverified")
        srs = dh.get("sector_relevance_score", 0)

        # ── Directors list (names only, comma-separated) ──────────────────────
        directors = c.get("directors") or []
        dir_names = ", ".join(
            d["name"].title() for d in directors
            if d.get("name") and not d.get("resigned_on")
        )[:120]

        # ── Address ───────────────────────────────────────────────────────────
        town   = c.get("town", "")
        county = c.get("county", "")

        # ── Col 1: Rank ───────────────────────────────────────────────────────
        cell(ws, row, 1, i, bg=bg, align="center", bold=True)

        # ── Col 2: Company name ───────────────────────────────────────────────
        cx_name = ws.cell(row=row, column=2, value=c["company_name"])
        cx_name.fill      = score_fill(acq)
        cx_name.font      = Font(name="Arial", size=10, bold=True,
                                 color=score_font_color(acq))
        cx_name.alignment = Alignment(horizontal="left", vertical="center")
        cx_name.border    = THIN

        # ── Col 3: Website URL (clickable style if live) ──────────────────────
        if website and website_live:
            cx_web = ws.cell(row=row, column=3, value=website)
            cx_web.font      = Font(name="Arial", size=8, color="0563C1", underline="single")
            cx_web.alignment = Alignment(horizontal="left", vertical="center")
            cx_web.border    = THIN
            if bg:
                if bg: cx_web.fill = bg if isinstance(bg, PatternFill) else fill(bg)
        else:
            cell(ws, row, 3, website or "—", bg=bg, fg="AAAAAA" if not website_live else None, size=8)

        # ── Col 4: Acquisition grade ──────────────────────────────────────────
        cx_grade = ws.cell(row=row, column=4, value=grade)
        cx_grade.fill      = score_fill(acq)
        cx_grade.font      = Font(name="Arial", size=9, bold=True,
                                  color=score_font_color(acq))
        cx_grade.alignment = Alignment(horizontal="center", vertical="center")
        cx_grade.border    = THIN

        # ── Col 5: Revenue estimate ───────────────────────────────────────────
        cell(ws, row, 5, rev_str, bg=bg, align="right", size=9)

        # ── Col 6: Employees ──────────────────────────────────────────────────
        emp_src = c.get("estimated_employees_source", "")
        emp_bg  = fill("E2EFDA") if emp_src.startswith("Tier 1") else (
                  fill("FFF2CC") if emp else bg)
        cell(ws, row, 6, emp if emp is not None else "—",
             bg=emp_bg, align="center", size=9, bold=emp_src.startswith("Tier 1"))

        # ── Col 7: About (web description) ────────────────────────────────────
        about_bg = fill("EBF5FB") if dh.get("website_description") else bg
        cell(ws, row, 7, about, bg=about_bg, size=8, wrap=True)

        # ── Col 8: Services (SIC) ─────────────────────────────────────────────
        cell(ws, row, 8, services_text, bg=bg, size=8, wrap=True)

        # ── Col 9: Sector match ───────────────────────────────────────────────
        cx_sm = ws.cell(row=row, column=9, value=srl)
        cx_sm.fill      = sector_fill_map.get(srl, fill("EEEEEE"))
        cx_sm.font      = Font(name="Arial", size=8, bold=True)
        cx_sm.alignment = Alignment(horizontal="center", vertical="center")
        cx_sm.border    = THIN
        if srs:
            cx_sm.comment = Comment(
                f"Sector relevance score: {srs}/100\n"
                + "\n".join(dh.get("sector_match_signals", [])),
                "V² Overview")

        # ── Col 10: Domain age ────────────────────────────────────────────────
        da = dh.get("domain_age_years")
        da_str = f"{da:.1f} yrs" if da else "—"
        da_bg  = fill("E2EFDA") if (da or 0) >= 10 else (
                 fill("FFF2CC") if (da or 0) >= 3 else bg)
        cell(ws, row, 10, da_str, bg=da_bg, align="center", size=8)

        # ── Col 11: LinkedIn ──────────────────────────────────────────────────
        li  = dh.get("has_linkedin", False)
        li_url = dh.get("linkedin_url") or contacts.get("linkedin_url") or ""
        if li_url:
            cx_li = ws.cell(row=row, column=11, value="✓ View")
            cx_li.font      = Font(name="Arial", size=8, color="0563C1", underline="single")
            cx_li.alignment = Alignment(horizontal="center", vertical="center")
            cx_li.border    = THIN
            cx_li.fill      = fill("E2EFDA")
        else:
            cell(ws, row, 11, "✓" if li else "—",
                 bg=fill("E2EFDA") if li else bg, align="center", size=9)

        # ── Col 12: Job postings ──────────────────────────────────────────────
        jobs = dh.get("has_job_postings", False)
        cell(ws, row, 12, "✓" if jobs else "—",
             bg=fill("E2EFDA") if jobs else bg, align="center", size=9)

        # ── Cols 13–16: Location + incorporation + directors ──────────────────
        cell(ws, row, 13, town,   bg=bg, size=8)
        cell(ws, row, 14, county, bg=bg, size=8)
        cell(ws, row, 15, (c.get("date_of_creation") or "")[:7], bg=bg, align="center", size=8)
        cell(ws, row, 16, dir_names, bg=bg, size=8, wrap=True)

        # ── Col 17: CH link ───────────────────────────────────────────────────
        ch_url = c.get("ch_url", "")
        if ch_url:
            cx_ch = ws.cell(row=row, column=17, value="View on CH")
            cx_ch.font      = Font(name="Arial", size=8, color="0563C1", underline="single")
            cx_ch.alignment = Alignment(horizontal="center", vertical="center")
            cx_ch.border    = THIN
            if bg:
                if bg: cx_ch.fill = bg if isinstance(bg, PatternFill) else fill(bg)
        else:
            cell(ws, row, 17, "—", bg=bg, align="center", size=8)

        # Row height — taller to show wrapped About text
        ws.row_dimensions[row].height = max(30, min(90,
            15 + (len(about) // 80) * 12))

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(companies)+3}"


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    enriched_path = os.path.join(cfg.OUTPUT_DIR, cfg.ENRICHED_JSON)
    with open(enriched_path) as f:
        companies = json.load(f)

    # Normalise all company records to a consistent schema before building
    # sheets.  This fills in computed fields (succession, sell_intent,
    # acquisition_score, estimated_employees …) for records that came from
    # enrich_batch.py (sector OCR runs) which only carry raw CH data.
    companies = [_normalise(c) for c in companies]

    # Re-sort by acquisition_score now that every company has one
    companies.sort(key=lambda x: x.get("acquisition_score", 0), reverse=True)

    bolt_on_path = os.path.join(cfg.OUTPUT_DIR, "bolt_on_analysis.json")
    bolt_on_data = {}
    if os.path.exists(bolt_on_path):
        with open(bolt_on_path) as f:
            bolt_on_data = json.load(f)

    wb = Workbook()
    build_pipeline(wb, companies)
    build_top30(wb, companies)
    build_overview(wb, companies)
    build_contacts(wb, companies)
    build_financials(wb, companies)
    if bolt_on_data:
        build_bolt_on(wb, bolt_on_data)
    build_sell_signals(wb, companies)
    build_contracts(wb, companies)
    build_digital_health(wb, companies)
    build_regulatory(wb, companies)
    build_competitors(wb, companies)
    build_summary(wb, companies)

    out_path = os.path.join(cfg.OUTPUT_DIR, cfg.EXCEL_OUTPUT)
    wb.save(out_path)
    print(f"Saved → {out_path}")
    return out_path


if __name__ == "__main__":
    run()
